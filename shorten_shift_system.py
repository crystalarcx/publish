# 互動式員工班表加班時數統計系統 (精簡優化版)
import pandas as pd
from datetime import datetime, date, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import streamlit as st
import warnings
import calendar
import io
import re
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass

warnings.filterwarnings('ignore')

# ===== 配置與資料類別 =====
@dataclass
class Config:
    DEFAULT_SHIFT_URL = "https://docs.google.com/spreadsheets/d/1JfhGZYRBWj6yp90o-sA0DrhzkcEM1Wfd_vqiEZEYd5c/edit?usp=sharing"
    DEFAULT_MAIN_URL = "https://docs.google.com/spreadsheets/d/1U8qLraVCRKJhySk0y93I_POP_LsgYjuS/edit?usp=sharing&ouid=115340390490868262616&rtpof=true&sd=true"
    OVERTIME_FORM_URL = "https://docs.google.com/document/d/1T75rw_3hQtIaBTGMFxa09G93Atihf4h-883Kg1tqPpo/edit?usp=sharing"
    ALLOWED_PERSONNEL = ['A30825', 'A408J6', 'A40837', 'A608Q2', 'A50847', 'A60811', 'A708J6', 'A808L5', 'B00505', 'A81205', 'A908H8']
    MAX_WEEKDAY_HOURS = 46.0
    AUTO_ADD_HOURS = 2.0
    WEEKEND_MIN_THRESHOLD = 3.0

@dataclass
class QueryResult:
    personnel: str
    year: int
    month: int
    daily_breakdown: Dict[str, float]
    weekday_hours: float
    weekend_hours: float
    total_hours: float

# ===== 核心功能類別 =====
class DataManager:
    @staticmethod
    def convert_url(url: str) -> str:
        """轉換Google Sheets URL為CSV格式"""
        if '/d/' not in url:
            return None
        try:
            sheet_id = url.split('/d/')[1].split('/')[0]
            return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
        except:
            return None

    @staticmethod
    @st.cache_data(ttl=300)
    def load_data(main_url: str, cache_version: int = 0) -> Tuple[pd.DataFrame, Dict, str]:
        """載入並處理資料"""
        try:
            progress = st.progress(0)
            status = st.empty()
            
            # 載入主要班表
            status.text("📊 載入班表資料...")
            progress.progress(30)
            main_csv = DataManager.convert_url(main_url)
            df = pd.read_csv(main_csv).iloc[:36, :83]
            
            # 載入班種對照表
            status.text("🔢 載入班種資料...")
            progress.progress(60)
            shift_csv = DataManager.convert_url(Config.DEFAULT_SHIFT_URL)
            shift_df = pd.read_csv(shift_csv)
            
            # 建立班種字典
            status.text("🔨 建立字典...")
            progress.progress(90)
            shift_dict = {}
            for _, row in shift_df.iterrows():
                if pd.notna(row.iloc[0]):
                    shift_dict[str(row.iloc[0]).strip()] = {
                        'overtime1': row.iloc[1] if len(row) > 1 else None,
                        'overtime2': row.iloc[2] if len(row) > 2 else None,
                        'cross_day': row.iloc[3] if len(row) > 3 else None
                    }
            
            progress.progress(100)
            progress.empty()
            status.empty()
            
            personnel_count = sum(1 for num in df.iloc[1, :] 
                                if pd.notna(num) and str(num).strip() in Config.ALLOWED_PERSONNEL)
            
            return df, shift_dict, f"✅ 載入成功！指定人員: {personnel_count} 人"
        except Exception as e:
            return None, None, f"❌ 載入失敗: {str(e)}"

class TimeCalculator:
    @staticmethod
    def calculate_hours(time_range) -> Optional[float]:
        """計算時間範圍的小時數"""
        if not time_range or pd.isna(time_range):
            return None
        
        time_str = str(time_range).strip()
        
        # 處理純數字
        if '-' not in time_str:
            try:
                return float(time_str.replace(',', '.'))
            except:
                return None
        
        # 處理時間範圍
        try:
            parts = time_str.replace(' ', '').split('-')
            if len(parts) != 2:
                return None
            
            start_time = TimeCalculator._parse_time(parts[0])
            end_time = TimeCalculator._parse_time(parts[1])
            
            if start_time is None or end_time is None:
                return None
            
            # 計算時差（處理跨日）
            if end_time <= start_time:
                end_time += 24 * 60
            
            return (end_time - start_time) / 60
        except:
            return None
    
    @staticmethod
    def _parse_time(time_str: str) -> Optional[int]:
        """解析時間為分鐘數"""
        time_str = time_str.strip()
        
        # HH:MM格式
        if ':' in time_str:
            try:
                parts = time_str.split(':')
                return int(parts[0]) * 60 + int(parts[1])
            except:
                pass
        
        # HHMM格式
        if len(time_str) == 4 and time_str.isdigit():
            try:
                return int(time_str[:2]) * 60 + int(time_str[2:])
            except:
                pass
        
        # HH格式
        if time_str.isdigit():
            try:
                return int(time_str) * 60
            except:
                pass
        
        return None

class OvertimeCalculator:
    @staticmethod
    def calculate_summary(personnel: str, year: int, month: int, df: pd.DataFrame, shift_dict: Dict) -> QueryResult:
        """計算加班時數統計"""
        # 找到匹配欄位
        personnel_row = df.iloc[1, :]
        matching_cols = [i for i, num in enumerate(personnel_row) 
                        if pd.notna(num) and str(num).strip() == personnel]
        
        if not matching_cols:
            return None
        
        daily_overtime = defaultdict(float)
        cross_day_records = defaultdict(float)
        worked_weekdays = set()
        
        # 處理每一天
        for day in range(1, calendar.monthrange(year, month)[1] + 1):
            try:
                current_date = date(year, month, day)
                date_str = f"{year}/{month:02d}/{day:02d}"
                is_weekend = OvertimeCalculator._is_weekend(current_date)
                
                # 處理所有匹配欄位
                for col_idx in matching_cols:
                    row_idx = day + 2
                    if row_idx >= len(df):
                        continue
                    
                    shift_value = str(df.iloc[row_idx, col_idx]).strip()
                    
                    if shift_value and shift_value != 'nan' and shift_value in shift_dict:
                        if not is_weekend:
                            worked_weekdays.add(date_str)
                        
                        shift_info = shift_dict[shift_value]
                        
                        # 計算當天加班時數
                        current_overtime = 0
                        for overtime_key in ['overtime1', 'overtime2']:
                            if pd.notna(shift_info[overtime_key]):
                                hours = TimeCalculator.calculate_hours(shift_info[overtime_key])
                                if hours:
                                    current_overtime += hours
                        
                        if current_overtime > 0:
                            daily_overtime[date_str] += current_overtime
                        
                        # 處理跨天時數
                        if pd.notna(shift_info['cross_day']):
                            cross_hours = TimeCalculator.calculate_hours(shift_info['cross_day'])
                            if cross_hours:
                                next_date = current_date + timedelta(days=1)
                                next_date_str = f"{next_date.year}/{next_date.month:02d}/{next_date.day:02d}"
                                daily_overtime[next_date_str] += cross_hours
            except:
                continue
        
        # 計算平日和假日時數
        weekday_hours = weekend_hours = 0
        for date_str, hours in daily_overtime.items():
            try:
                parts = date_str.split('/')
                check_date = date(int(parts[0]), int(parts[1]), int(parts[2]))
                if OvertimeCalculator._is_weekend(check_date):
                    weekend_hours += hours
                else:
                    weekday_hours += hours
            except:
                continue
        
        # 調整平日時數（46小時邏輯）
        weekday_hours, daily_overtime = OvertimeCalculator._adjust_weekday_hours(
            weekday_hours, daily_overtime, worked_weekdays, year, month
        )
        
        return QueryResult(
            personnel=personnel,
            year=year,
            month=month,
            daily_breakdown=dict(daily_overtime),
            weekday_hours=weekday_hours,
            weekend_hours=weekend_hours,
            total_hours=weekday_hours + weekend_hours
        )
    
    @staticmethod
    def _is_weekend(check_date: date) -> bool:
        """檢查是否為週末或自定義假日"""
        date_key = f"{check_date.year}-{check_date.month:02d}-{check_date.day:02d}"
        if 'custom_holidays' in st.session_state and date_key in st.session_state.custom_holidays:
            return True
        return check_date.weekday() >= 5
    
    @staticmethod
    def _adjust_weekday_hours(weekday_hours: float, daily_overtime: dict, worked_weekdays: set, year: int, month: int) -> Tuple[float, dict]:
        """調整平日時數到46小時"""
        daily_overtime = defaultdict(float, daily_overtime)
        
        if weekday_hours > Config.MAX_WEEKDAY_HOURS:
            # 減少超過部分
            excess = weekday_hours - Config.MAX_WEEKDAY_HOURS
            weekday_dates = [(date_str, hours) for date_str, hours in daily_overtime.items() 
                           if hours > 0 and not OvertimeCalculator._is_date_weekend(date_str)]
            weekday_dates.sort(key=lambda x: x[1])
            
            removed = 0
            for date_str, hours in weekday_dates:
                if removed + hours <= excess:
                    daily_overtime[date_str] = 0
                    removed += hours
                    weekday_hours -= hours
                elif removed < excess:
                    reduce_amount = excess - removed
                    daily_overtime[date_str] -= reduce_amount
                    weekday_hours -= reduce_amount
                    break
        
        elif weekday_hours < Config.MAX_WEEKDAY_HOURS:
            # 補足到46小時
            shortage = Config.MAX_WEEKDAY_HOURS - weekday_hours
            available_days = []
            
            for day in range(1, calendar.monthrange(year, month)[1] + 1):
                try:
                    check_date = date(year, month, day)
                    date_str = f"{year}/{month:02d}/{day:02d}"
                    if not OvertimeCalculator._is_weekend(check_date) and date_str not in worked_weekdays:
                        priority = 1 if check_date.weekday() in [1, 3] else 2  # 週二四優先
                        available_days.append((date_str, priority))
                except:
                    continue
            
            available_days.sort(key=lambda x: x[1])
            days_needed = int(shortage / Config.AUTO_ADD_HOURS) + (1 if shortage % Config.AUTO_ADD_HOURS > 0 else 0)
            
            for i, (date_str, _) in enumerate(available_days[:days_needed]):
                daily_overtime[date_str] += Config.AUTO_ADD_HOURS
                weekday_hours += Config.AUTO_ADD_HOURS
        
        return weekday_hours, daily_overtime
    
    @staticmethod
    def _is_date_weekend(date_str: str) -> bool:
        """檢查日期字串是否為週末"""
        try:
            parts = date_str.split('/')
            check_date = date(int(parts[0]), int(parts[1]), int(parts[2]))
            return OvertimeCalculator._is_weekend(check_date)
        except:
            return False

class ExcelExporter:
    @staticmethod
    def export_excel(result: QueryResult, df: pd.DataFrame, shift_dict: Dict) -> Tuple[bool, Any, float, float, float, int]:
        """匯出Excel報表"""
        try:
            # 收集時間字串和建立Excel資料
            excel_data = ExcelExporter._build_excel_data(result, df, shift_dict)
            
            # 創建Excel檔案
            output = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{result.personnel}加班統計"
            
            # 設定標題行
            headers = ['日期', '原始時間字串', '平日時數', '假日時數', '工作類型']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # 填入資料
            for row_idx, data in enumerate(excel_data, 2):
                for col_idx, value in enumerate([data['日期'], data['時間字串'], 
                                               data['平日時數'], data['假日時數'], data['工作類型']], 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 設定欄寬和統計
            for col_idx, width in enumerate([8, 30, 12, 12, 15], 1):
                ws.column_dimensions[chr(64 + col_idx)].width = width
            
            total_weekday = sum(data['平日時數'] for data in excel_data)
            total_weekend = sum(data['假日時數'] for data in excel_data)
            
            wb.save(output)
            output.seek(0)
            
            return True, output, total_weekday, total_weekend, total_weekday + total_weekend, len(excel_data)
        except Exception as e:
            return False, f"Excel匯出失敗: {str(e)}", 0, 0, 0, 0
    
    @staticmethod
    def _build_excel_data(result: QueryResult, df: pd.DataFrame, shift_dict: Dict) -> List[Dict]:
        """建立Excel資料"""
        excel_data = []
        
        for day in range(1, calendar.monthrange(result.year, result.month)[1] + 1):
            date_str = f"{result.year}/{result.month:02d}/{day:02d}"
            
            if date_str in result.daily_breakdown:
                hours = result.daily_breakdown[date_str]
                is_weekend = OvertimeCalculator._is_date_weekend(date_str)
                
                # 簡化的時間字串收集
                time_strings = ExcelExporter._get_time_strings(df, shift_dict, result.personnel, day, result.year, result.month)
                
                weekday_hours = 0 if is_weekend else hours
                weekend_hours = hours if is_weekend else 0
                
                # 假日邏輯調整
                if is_weekend and weekend_hours <= Config.WEEKEND_MIN_THRESHOLD and weekend_hours > 0:
                    time_strings = "12:00-14:00(撰寫病歷)," + time_strings if time_strings else "12:00-14:00(撰寫病歷)"
                    weekend_hours += Config.AUTO_ADD_HOURS
                
                work_type = ExcelExporter._extract_work_type(time_strings)
                
                excel_data.append({
                    '日期': f"{day:02d}",
                    '時間字串': time_strings,
                    '平日時數': weekday_hours,
                    '假日時數': weekend_hours,
                    '工作類型': work_type
                })
        
        return excel_data
    
    @staticmethod
    def _get_time_strings(df: pd.DataFrame, shift_dict: Dict, personnel: str, day: int, year: int, month: int) -> str:
        """獲取時間字串"""
        personnel_row = df.iloc[1, :]
        matching_cols = [i for i, num in enumerate(personnel_row) 
                        if pd.notna(num) and str(num).strip() == personnel]
        
        time_strings = []
        for col_idx in matching_cols:
            row_idx = day + 2
            if row_idx < len(df):
                shift_value = str(df.iloc[row_idx, col_idx]).strip()
                if shift_value in shift_dict:
                    shift_info = shift_dict[shift_value]
                    for key in ['overtime1', 'overtime2']:
                        if pd.notna(shift_info[key]):
                            time_strings.append(str(shift_info[key]).strip())
        
        return ",".join(time_strings)
    
    @staticmethod
    def _extract_work_type(time_string: str) -> str:
        """提取工作類型"""
        if not time_string:
            return "會議"
        chinese_match = re.search(r'[\u4e00-\u9fff]+', time_string)
        return chinese_match.group() if chinese_match else "臨床業務"

# ===== Streamlit 界面 =====
def main():
    st.set_page_config(page_title="員工班表系統", page_icon="🏢", layout="wide")
    
    # 初始化session state
    for key in ['df', 'shift_dict', 'custom_holidays', 'last_result', 'current_page']:
        if key not in st.session_state:
            st.session_state[key] = None if key != 'custom_holidays' else {}
    
    if 'current_page' not in st.session_state:
        st.session_state.current_page = "載入資料"
    
    st.title("🏢 員工班表加班時數統計系統")
    
    # 側邊欄導航
    with st.sidebar:
        st.header("📋 功能選單")
        
        pages = ["載入資料", "查詢統計", "假日管理"]
        for page in pages:
            if st.button(page, type="primary" if st.session_state.current_page == page else "secondary"):
                st.session_state.current_page = page
                st.rerun()
        
        st.markdown("---")
        st.markdown(f"🔗 [空白加班單]({Config.OVERTIME_FORM_URL})")
        
        if st.button("🗑️ 清除快取"):
            st.cache_data.clear()
            st.success("快取已清除")
    
    # 主要內容區域
    if st.session_state.current_page == "載入資料":
        load_data_page()
    elif st.session_state.current_page == "查詢統計":
        query_page()
    elif st.session_state.current_page == "假日管理":
        holiday_page()

def load_data_page():
    st.header("📥 載入班表資料")
    
    with st.form("load_form"):
        url = st.text_area("Google Sheets 連結", placeholder="請貼上班表連結")
        
        col1, col2 = st.columns(2)
        with col1:
            submit = st.form_submit_button("📥 載入資料", type="primary")
        with col2:
            default = st.form_submit_button("🔄 載入預設")
    
    if default:
        url = Config.DEFAULT_MAIN_URL
        submit = True
    
    if submit:
        if not url.strip():
            st.error("請輸入連結")
            return
        
        df, shift_dict, message = DataManager.load_data(url)
        
        if df is not None:
            st.session_state.df = df
            st.session_state.shift_dict = shift_dict
            st.success(message)
            
            with st.expander("資料預覽"):
                st.dataframe(df.head())
        else:
            st.error(message)

def query_page():
    st.header("🔍 加班時數查詢")
    
    if st.session_state.df is None:
        st.warning("請先載入資料")
        return
    
    df = st.session_state.df
    personnel_options = [f"{num} (Column {chr(65+i)})" for i, num in enumerate(df.iloc[1, :]) 
                        if pd.notna(num) and str(num).strip() in Config.ALLOWED_PERSONNEL]
    
    if not personnel_options:
        st.error("未找到指定人事號")
        return
    
    with st.form("query_form"):
        col1, col2, col3 = st.columns(3)
        with col1:
            personnel = st.selectbox("人事號", personnel_options)
        with col2:
            year = st.number_input("年份", min_value=2020, max_value=2030, value=datetime.now().year)
        with col3:
            month = st.selectbox("月份", range(1, 13), index=datetime.now().month-1)
        
        submit = st.form_submit_button("🔍 查詢", type="primary")
    
    if submit:
        target_personnel = personnel.split(' (')[0]
        
        with st.spinner("查詢中..."):
            result = OvertimeCalculator.calculate_summary(
                target_personnel, year, month+1, df, st.session_state.shift_dict
            )
        
        if result:
            st.session_state.last_result = result
            
            # 顯示結果
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("平日時數", f"{result.weekday_hours:.1f}h")
            with col2:
                st.metric("假日時數", f"{result.weekend_hours:.1f}h")
            with col3:
                st.metric("總時數", f"{result.total_hours:.1f}h")
            
            # 詳細資料
            if result.daily_breakdown:
                st.subheader("每日明細")
                breakdown_data = []
                for date_str, hours in sorted(result.daily_breakdown.items()):
                    if hours > 0:
                        is_weekend = OvertimeCalculator._is_date_weekend(date_str)
                        breakdown_data.append({
                            '日期': date_str,
                            '時數': f"{hours:.1f}h",
                            '類型': '假日' if is_weekend else '平日'
                        })
                
                if breakdown_data:
                    st.dataframe(pd.DataFrame(breakdown_data), use_container_width=True)
        else:
            st.error("查詢失敗")
    
    # Excel匯出
    if st.session_state.last_result:
        st.subheader("📊 匯出報表")
        if st.button("產生Excel"):
            with st.spinner("生成中..."):
                success, content, wd, we, total, rows = ExcelExporter.export_excel(
                    st.session_state.last_result, df, st.session_state.shift_dict
                )
            
            if success:
                st.success("Excel生成成功！")
                filename = f"{st.session_state.last_result.personnel}_{st.session_state.last_result.year}年{st.session_state.last_result.month:02d}月_加班統計.xlsx"
                st.download_button("📥 下載", content.getvalue(), filename,
                                 "application/vnd.openxmlformats-officeedocument.spreadsheetml.sheet")
            else:
                st.error(content)

def holiday_page():
    st.header("🗓️ 假日管理")
    st.warning("⚠️ 假日設定在關閉瀏覽器後會清除")
    
    with st.form("holiday_form"):
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            year = st.number_input("年", min_value=2020, max_value=2030, value=datetime.now().year)
        with col2:
            month = st.selectbox("月", range(1, 13), index=datetime.now().month-1)
        with col3:
            day = st.number_input("日", min_value=1, max_value=31, value=1)
        with col4:
            reason = st.text_input("原因", value="自定義假日")
        
        col_add, col_remove = st.columns(2)
        with col_add:
            add = st.form_submit_button("➕ 新增", type="primary")
        with col_remove:
            remove = st.form_submit_button("❌ 移除")
    
    if add:
        try:
            test_date = date(year, month+1, day)
            date_key = f"{year}-{month+1:02d}-{day:02d}"
            weekday = ['一', '二', '三', '四', '五', '六', '日'][test_date.weekday()]
            
            st.session_state.custom_holidays[date_key] = f"{reason}({weekday})"
            st.success(f"已新增: {date_key} {reason}({weekday})")
            st.rerun()
        except ValueError:
            st.error("無效日期")
    
    if remove:
        date_key = f"{year}-{month+1:02d}-{day:02d}"
        if date_key in st.session_state.custom_holidays:
            removed = st.session_state.custom_holidays.pop(date_key)
            st.success(f"已移除: {date_key} ({removed})")
            st.rerun()
        else:
            st.warning("該日期不是自定義假日")
    
    # 顯示現有假日
    if st.session_state.custom_holidays:
        st.subheader(f"目前假日 ({len(st.session_state.custom_holidays)} 天)")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🗑️ 清除全部"):
                st.session_state.custom_holidays.clear()
                st.success("已清除所有假日")
                st.rerun()
        
        holiday_list = []
        for date_key, desc in sorted(st.session_state.custom_holidays.items()):
            holiday_list.append({'日期': date_key, '描述': desc})
        
        st.dataframe(pd.DataFrame(holiday_list), use_container_width=True)
    else:
        st.info("目前無自定義假日")

if __name__ == "__main__":
    main()