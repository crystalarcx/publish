# 互動式員工班表加班時數統計系統 (Streamlit版)
import pandas as pd
from datetime import datetime, date, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import streamlit as st
import warnings
import calendar
import io
import base64
warnings.filterwarnings('ignore')

# ===== Streamlit 頁面配置 =====
st.set_page_config(
    page_title="員工班表加班時數統計系統",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ===== 初始化 Session State =====
if 'df' not in st.session_state:
    st.session_state.df = None
if 'shift_dict' not in st.session_state:
    st.session_state.shift_dict = {}
if 'custom_holidays' not in st.session_state:
    st.session_state.custom_holidays = {}
if 'version_info' not in st.session_state:
    st.session_state.version_info = {
        'version_name': '',
        'main_sheet_url': '',
        'shift_sheet_url': '',
        'load_time': ''
    }
if 'last_query_result' not in st.session_state:
    st.session_state.last_query_result = None
if 'current_page' not in st.session_state:
    st.session_state.current_page = "載入班表資料"

# ===== 工具函數 =====
def convert_google_sheet_url(url):
    """將 Google Sheets URL 轉換為可直接讀取的 CSV URL"""
    if '/d/' in url:
        sheet_id = url.split('/d/')[1].split('/')[0]
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    return None

@st.cache_data
def load_data_from_urls(main_sheet_url, version_name):
    """使用自定義URL載入資料"""
    # 班種對照表固定使用預設連結
    default_shift_sheet_url = "https://docs.google.com/spreadsheets/d/1JfhGZYRBWj6yp90o-sA0DrhzkcEM1Wfd_vqiEZEYd5c/edit?usp=sharing"

    main_csv_url = convert_google_sheet_url(main_sheet_url)
    shift_csv_url = convert_google_sheet_url(default_shift_sheet_url)
    
    if not main_csv_url or not shift_csv_url:
        return None, None, "❌ 無效的Google Sheets URL格式"

    try:
        # 讀取員工班表
        df_full = pd.read_csv(main_csv_url)
        df = df_full.iloc[:36, :83]  # 選取 A1:CE36 範圍

        # 讀取班種對照表
        shift_df = pd.read_csv(shift_csv_url)

        # 建立班種字典
        shift_dict = {}
        for index, row in shift_df.iterrows():
            shift_type = str(row.iloc[0]).strip()
            overtime_hours_1 = row.iloc[1]
            overtime_hours_2 = row.iloc[2]
            cross_day_hours = row.iloc[3] if len(row) > 3 else None

            shift_dict[shift_type] = {
                'overtime_hours_1': overtime_hours_1,
                'overtime_hours_2': overtime_hours_2,
                'cross_day_hours': cross_day_hours
            }

        return df, shift_dict, f"✅ 資料讀取成功！版本: {version_name}, 班表: {df.shape}, 班種: {len(shift_dict)} 種"
    except Exception as e:
        return None, None, f"❌ 資料讀取失敗: {e}"

def calculate_hours(time_range):
    """計算時間範圍的小時數"""
    try:
        if not time_range or pd.isna(time_range):
            return None

        time_str = str(time_range).strip()

        # 處理逗號作為小數點的情況
        if ',' in time_str and '-' not in time_str:
            try:
                hours = float(time_str.replace(',', '.'))
                return hours
            except:
                pass

        # 如果不包含分隔符，可能是單純的小時數
        if '-' not in time_str:
            try:
                hours = float(time_str)
                return hours
            except:
                return None

        # 清理時間字串
        time_str = time_str.replace(' ', '').replace(',', '')

        if '-' in time_str:
            parts = time_str.split('-')
            if len(parts) != 2:
                return None
            start_str, end_str = parts
        else:
            return None

        def parse_time_hhmm(time_str):
            """解析 HH:MM 格式的時間"""
            time_str = time_str.strip()

            # 處理 HH:MM 格式
            if ':' in time_str:
                try:
                    parts = time_str.split(':')
                    if len(parts) == 2:
                        hour = int(parts[0])
                        minute = int(parts[1])
                        if 0 <= hour <= 23 and 0 <= minute <= 59:
                            return hour, minute
                except ValueError:
                    pass

            # 處理 HHMM 格式
            if len(time_str) == 4 and time_str.isdigit():
                try:
                    hour = int(time_str[:2])
                    minute = int(time_str[2:])
                    if 0 <= hour <= 23 and 0 <= minute <= 59:
                        return hour, minute
                except ValueError:
                    pass

            # 處理 HH 格式
            if time_str.isdigit() and 1 <= len(time_str) <= 2:
                try:
                    hour = int(time_str)
                    if 0 <= hour <= 23:
                        return hour, 0
                except ValueError:
                    pass

            # 處理帶小數點的小時數格式
            try:
                hour_decimal = float(time_str)
                if 0 <= hour_decimal <= 24:
                    hour = int(hour_decimal)
                    minute = int((hour_decimal - hour) * 60)
                    if 0 <= hour <= 23 and 0 <= minute <= 59:
                        return hour, minute
            except ValueError:
                pass

            return None, None

        # 解析開始和結束時間
        start_hour, start_min = parse_time_hhmm(start_str)
        end_hour, end_min = parse_time_hhmm(end_str)

        if start_hour is None or end_hour is None:
            return None

        # 轉換為分鐘
        start_minutes = start_hour * 60 + start_min
        end_minutes = end_hour * 60 + end_min

        # 處理跨日情況
        if end_minutes <= start_minutes:
            end_minutes += 24 * 60

        # 計算時數
        total_minutes = end_minutes - start_minutes
        hours = total_minutes / 60

        return hours

    except Exception as e:
        return None

def get_day_type(year, month, day):
    """判斷日期是平日還是假日（含自定義假日）"""
    try:
        # 檢查是否為自定義假日
        date_key = f"{year}-{month:02d}-{day:02d}"
        if date_key in st.session_state.custom_holidays:
            return st.session_state.custom_holidays[date_key], True

        # 一般週末判斷
        current_date = date(year, month, day)
        weekday = current_date.weekday()  # 0=星期一, 6=星期日

        if weekday == 5:  # 星期六
            return "假日(六)", True
        elif weekday == 6:  # 星期日
            return "假日(日)", True
        else:  # 星期一到星期五
            weekdays = ["一", "二", "三", "四", "五"]
            return f"平日({weekdays[weekday]})", False
    except ValueError:
        return "無效日期", False

def extract_chinese_note(time_string):
    """從時間字串中提取中文註記"""
    import re

    if not time_string:
        return "臨床業務"

    chinese_pattern = r'[\u4e00-\u9fff]+|\([^\)]*[\u4e00-\u9fff][^\)]*\)'
    chinese_matches = re.findall(chinese_pattern, time_string)

    if chinese_matches:
        chinese_note = chinese_matches[0]
        chinese_note = chinese_note.replace('(', '').replace(')', '')
        return chinese_note
    else:
        return "臨床業務"

def calculate_overtime_summary(target_personnel, year, month, matching_columns):
    """計算指定人員的加班時數統計"""
    df = st.session_state.df
    shift_dict = st.session_state.shift_dict
    
    weekday_overtime = 0.0
    weekend_overtime = 0.0
    total_overtime = 0.0
    daily_records = []
    cross_day_records = defaultdict(float)
    worked_weekdays = set()

    # 對每個匹配的欄位進行處理
    for col_idx in matching_columns:
        column_data = df.iloc[:, col_idx]

        # 處理每一天
        for day in range(1, 32):
            try:
                current_date = date(year, month, day)
            except ValueError:
                continue

            row_idx = day + 2

            if row_idx < len(column_data):
                value = column_data.iloc[row_idx]
                date_str = f"{year}/{month:02d}/{day:02d}"
                day_type, is_weekend = get_day_type(year, month, day)

                shift_value = str(value).strip()

                # 記錄有上班的平日
                if shift_value and pd.notna(value) and shift_value != 'nan' and not is_weekend:
                    worked_weekdays.add(date_str)

                if shift_value in shift_dict and pd.notna(value) and shift_value:
                    shift_info = shift_dict[shift_value]
                    overtime_1 = shift_info['overtime_hours_1']
                    overtime_2 = shift_info['overtime_hours_2']
                    cross_day = shift_info['cross_day_hours']

                    current_day_overtime = 0.0
                    next_day_overtime = 0.0

                    # 計算當天加班時數
                    if pd.notna(overtime_1) and str(overtime_1).strip():
                        calculated_hours_1 = calculate_hours(str(overtime_1))
                        if calculated_hours_1 is not None:
                            current_day_overtime += calculated_hours_1

                    if pd.notna(overtime_2) and str(overtime_2).strip():
                        calculated_hours_2 = calculate_hours(str(overtime_2))
                        if calculated_hours_2 is not None:
                            current_day_overtime += calculated_hours_2

                    # 計算跨天時數
                    if pd.notna(cross_day) and str(cross_day).strip():
                        cross_day_hours = calculate_hours(str(cross_day))
                        if cross_day_hours is not None:
                            next_day_overtime = cross_day_hours
                            next_date = current_date + timedelta(days=1)
                            next_date_str = f"{next_date.year}/{next_date.month:02d}/{next_date.day:02d}"
                            cross_day_records[next_date_str] += next_day_overtime

                    # 記錄當天資料
                    if current_day_overtime > 0 or next_day_overtime > 0:
                        daily_records.append({
                            'date': date_str,
                            'day_type': day_type,
                            'is_weekend': is_weekend,
                            'shift': shift_value,
                            'current_day_overtime': current_day_overtime,
                            'cross_day_overtime': next_day_overtime
                        })

    # 建立完整的日期加班時數記錄
    final_daily_overtime = defaultdict(float)

    # 先加入當天的加班時數
    for record in daily_records:
        date_str = record['date']
        current_overtime = record['current_day_overtime']
        if current_overtime > 0:
            final_daily_overtime[date_str] += current_overtime

    # 再加入跨天時數
    for date_str, cross_hours in cross_day_records.items():
        final_daily_overtime[date_str] += cross_hours

    # 計算平日和假日總時數
    for date_str, total_hours in final_daily_overtime.items():
        try:
            date_parts = date_str.split('/')
            check_year = int(date_parts[0])
            check_month = int(date_parts[1])
            check_day = int(date_parts[2])

            day_type, is_weekend = get_day_type(check_year, check_month, check_day)

            if is_weekend:
                weekend_overtime += total_hours
            else:
                weekday_overtime += total_hours

            total_overtime += total_hours

        except (ValueError, IndexError):
            continue

    # 檢查平日加班時數是否小於46小時，如果是則自動補足
    if weekday_overtime < 46:
        shortage = 46 - weekday_overtime

        # 找出所有該月的平日且沒有上班的日期
        available_weekdays = []
        for day in range(1, 32):
            try:
                check_date = date(year, month, day)
                date_str = f"{year}/{month:02d}/{day:02d}"
                day_type, is_weekend = get_day_type(year, month, day)
                weekday_num = check_date.weekday()

                if not is_weekend and date_str not in worked_weekdays:
                    if weekday_num in [1, 3]:  # 週二、週四
                        priority = 1
                    elif weekday_num in [0, 2, 4]:  # 週一、週三、週五
                        priority = 2
                    else:
                        priority = 3

                    available_weekdays.append((date_str, day_type, weekday_num, priority))
            except ValueError:
                continue

        # 按優先順序排序
        available_weekdays.sort(key=lambda x: (x[3], x[0]))

        if available_weekdays:
            days_needed = int(shortage / 2) + (1 if shortage % 2 > 0 else 0)

            for i, (date_str, day_type, weekday_num, priority) in enumerate(available_weekdays):
                if i < days_needed:
                    final_daily_overtime[date_str] += 2.0
                    weekday_overtime += 2.0
                    total_overtime += 2.0

    return weekday_overtime, weekend_overtime, total_overtime, daily_records, final_daily_overtime

def export_to_excel(target_personnel, year, month, matching_columns, export_data):
    """導出Excel報表"""
    try:
        df = st.session_state.df
        shift_dict = st.session_state.shift_dict
        
        excel_data = []
        date_time_strings = defaultdict(list)

        # 收集原始時間字串
        for col_idx in matching_columns:
            column_data = df.iloc[:, col_idx]

            for day in range(1, 32):
                try:
                    current_date = date(year, month, day)
                    date_str = f"{year}/{month:02d}/{day:02d}"
                    row_idx = day + 2

                    if row_idx < len(column_data):
                        value = column_data.iloc[row_idx]
                        shift_value = str(value).strip()

                        if shift_value in shift_dict and pd.notna(value) and shift_value:
                            shift_info = shift_dict[shift_value]

                            current_day_time_strings = []

                            if pd.notna(shift_info['overtime_hours_1']) and str(shift_info['overtime_hours_1']).strip():
                                current_day_time_strings.append(str(shift_info['overtime_hours_1']).strip())

                            if pd.notna(shift_info['overtime_hours_2']) and str(shift_info['overtime_hours_2']).strip():
                                current_day_time_strings.append(str(shift_info['overtime_hours_2']).strip())

                            if current_day_time_strings:
                                date_time_strings[date_str].extend(current_day_time_strings)

                            if pd.notna(shift_info['cross_day_hours']) and str(shift_info['cross_day_hours']).strip():
                                cross_day_time_str = str(shift_info['cross_day_hours']).strip()
                                next_date = current_date + timedelta(days=1)
                                next_date_str = f"{next_date.year}/{next_date.month:02d}/{next_date.day:02d}"
                                date_time_strings[next_date_str].append(cross_day_time_str)

                except ValueError:
                    continue

        # 建立Excel資料
        for day in range(1, 32):
            try:
                current_date = date(year, month, day)
                date_str = f"{year}/{month:02d}/{day:02d}"
                day_type, is_weekend = get_day_type(year, month, day)

                time_strings = date_time_strings.get(date_str, [])
                original_time_str = ",".join(time_strings) if time_strings else ""

                weekday_hours = 0.0
                weekend_hours = 0.0

                if date_str in export_data:
                    total_hours = export_data[date_str]
                    if is_weekend:
                        weekend_hours = total_hours
                    else:
                        weekday_hours = total_hours

                work_type = ""
                if date_str in export_data and not original_time_str:
                    original_time_str = "14:00-16:00(會議)"
                    work_type = "會議"
                else:
                    work_type = extract_chinese_note(original_time_str)

                if original_time_str or weekday_hours > 0 or weekend_hours > 0:
                    excel_data.append({
                        '日期': f"{day:02d}",
                        '原始時間字串': original_time_str,
                        '平日時數': weekday_hours,
                        '假日時數': weekend_hours,
                        '工作類型': work_type
                    })

            except ValueError:
                continue

        # 建立Excel檔案
        df_excel = pd.DataFrame(excel_data)

        # 創建Excel內容到內存
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{target_personnel}加班統計"

        # 設定標題
        headers = ['日期', '原始時間字串', '平日時數', '假日時數', '工作類型']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', size=12)

        # 設定邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 填入資料
        for row_idx, row_data in enumerate(df_excel.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                if col_idx in [3, 4]:  # 平日時數、假日時數
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if value > 0:
                        cell.number_format = '0.0'
                elif col_idx == 5:  # 工作類型
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # 調整欄寬
        column_widths = [8, 30, 12, 12, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # 添加統計
        total_weekday = df_excel['平日時數'].sum()
        total_weekend = df_excel['假日時數'].sum()
        total_hours = total_weekday + total_weekend

        last_row = len(df_excel) + 3

        ws.cell(row=last_row, column=1, value="統計總計").font = Font(bold=True, size=12)
        ws.cell(row=last_row, column=1).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        ws.cell(row=last_row + 1, column=1, value="平日加班總時數:")
        ws.cell(row=last_row + 1, column=2, value=f"{total_weekday:.1f} 小時")

        ws.cell(row=last_row + 2, column=1, value="假日加班總時數:")
        ws.cell(row=last_row + 2, column=2, value=f"{total_weekend:.1f} 小時")

        ws.cell(row=last_row + 3, column=1, value="總加班時數:")
        ws.cell(row=last_row + 3, column=2, value=f"{total_hours:.1f} 小時")
        ws.cell(row=last_row + 3, column=2).font = Font(bold=True)

        wb.save(output)
        output.seek(0)

        return True, output, total_weekday, total_weekend, total_hours, len(df_excel)

    except Exception as e:
        return False, str(e), 0, 0, 0, 0

# ===== 主要界面 =====
def main():
    st.title("🏢 員工班表加班時數統計系統")
    
    # 側邊欄 - 改為直接按鈕選擇
    with st.sidebar:
        st.header("📋 系統功能")
        
        # 直接展開的按鈕選項
        if st.button("📥 載入班表資料", type="primary" if st.session_state.current_page == "載入班表資料" else "secondary"):
            st.session_state.current_page = "載入班表資料"
            st.experimental_rerun()
        
        if st.button("🔍 查詢加班時數", type="primary" if st.session_state.current_page == "查詢加班時數" else "secondary"):
            st.session_state.current_page = "查詢加班時數"
            st.experimental_rerun()
        
        if st.button("🗓️ 自定義假日管理", type="primary" if st.session_state.current_page == "自定義假日管理" else "secondary"):
            st.session_state.current_page = "自定義假日管理"
            st.experimental_rerun()
    
    # 顯示版本資訊
    if st.session_state.version_info['version_name']:
        st.success(f"📋 目前載入版本: **{st.session_state.version_info['version_name']}**")
        st.info(f"⏰ 載入時間: {st.session_state.version_info['load_time']}")
    else:
        st.warning("📋 尚未載入任何班表資料")
    
    # 根據當前頁面顯示對應內容
    if st.session_state.current_page == "載入班表資料":
        load_data_page()
    elif st.session_state.current_page == "查詢加班時數":
        query_page()
    elif st.session_state.current_page == "自定義假日管理":
        holiday_management_page()

def load_data_page():
    """載入資料頁面"""
    st.header("📥 載入雲端班表資料")
    
    with st.form("load_data_form"):
        version_name = st.text_input(
            "版本識別名稱",
            placeholder="請輸入版本識別名稱 (例如: 2024年1月班表)",
            help="為此次載入的班表資料命名，方便識別"
        )
        
        main_sheet_url = st.text_area(
            "員工班表 Google Sheets 連結",
            placeholder="請貼上員工班表的 Google Sheets 完整連結",
            help="請確保 Google Sheets 已設定為「知道連結的使用者」可檢視"
        )
        
        st.info("📋 班種對照表: 系統將自動使用預設的班種對照表")
        
        col1, col2 = st.columns(2)
        with col1:
            submit_button = st.form_submit_button("📥 載入雲端班表", type="primary")
        with col2:
            use_default = st.form_submit_button("🔗 使用預設班表")
    
    if use_default:
        version_name = "預設班表"
        main_sheet_url = "https://docs.google.com/spreadsheets/d/1U8qLraVCRKJhySk0y93I_POP_LsgYjuS/edit?usp=sharing&ouid=115340390490868262616&rtpof=true&sd=true"
        st.info("✅ 已填入預設班表連結")
        submit_button = True
    
    if submit_button:
        if not version_name.strip():
            st.error("❌ 請輸入版本識別名稱")
            return
        
        if not main_sheet_url.strip():
            st.error("❌ 請輸入員工班表的 Google Sheets 連結")
            return
        
        if '/d/' not in main_sheet_url:
            st.error("❌ Google Sheets 連結格式不正確，請確保包含完整的分享連結")
            return
        
        with st.spinner(f"🔄 正在載入「{version_name}」班表資料..."):
            df, shift_dict, message = load_data_from_urls(main_sheet_url, version_name)
        
        if df is not None:
            # 更新 session state
            st.session_state.df = df
            st.session_state.shift_dict = shift_dict
            st.session_state.version_info = {
                'version_name': version_name,
                'main_sheet_url': main_sheet_url,
                'shift_sheet_url': "系統預設",
                'load_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            # 取得人事號清單
            personnel_numbers = df.iloc[1, :].tolist()
            personnel_count = sum(1 for num in personnel_numbers if pd.notna(num) and str(num).strip())
            
            st.success(f"✅ 「{version_name}」載入完成！")
            st.info(f"📊 班表尺寸: {df.shape}")
            st.info(f"👥 找到 {personnel_count} 個人事號")
            st.info(f"🔢 班種數量: {len(shift_dict)} 種")
            
            # 清除快取
            st.cache_data.clear()
            
        else:
            st.error(message)

def query_page():
    """查詢頁面"""
    st.header("🔍 員工加班時數查詢")
    
    if st.session_state.df is None:
        st.warning("⚠️ 請先載入班表資料")
        return
    
    df = st.session_state.df
    
    # 取得人事號清單
    personnel_numbers = df.iloc[1, :].tolist()
    personnel_options = []
    
    for i, num in enumerate(personnel_numbers):
        if pd.notna(num) and str(num).strip():
            col_name = chr(65 + i) if i < 26 else chr(65 + i//26 - 1) + chr(65 + i%26)
            personnel_options.append(f"{num} (Column {col_name})")
    
    if not personnel_options:
        st.error("❌ 未找到任何人事號")
        return
    
    with st.form("query_form"):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            selected_personnel = st.selectbox("選擇人事號", personnel_options)
        
        with col2:
            year = st.number_input("西元年", min_value=2020, max_value=2030, 
                                 value=datetime.now().year)
        
        with col3:
            month = st.selectbox("月份", 
                               [(i, f"{i}月") for i in range(1, 13)], 
                               index=datetime.now().month-1,
                               format_func=lambda x: x[1])
        
        submit_query = st.form_submit_button("🔍 查詢加班時數", type="primary")
    
    if submit_query:
        # 提取人事號
        target_personnel = selected_personnel.split(' (')[0]
        month_value = month[0]
        
        with st.spinner(f"🔍 正在查詢 {target_personnel} 的 {year}年{month_value}月 加班時數..."):
            # 查找匹配的欄位
            personnel_numbers = df.iloc[1, :].tolist()
            matching_columns = []
            
            for col_idx, personnel_num in enumerate(personnel_numbers):
                if pd.notna(personnel_num) and str(personnel_num).strip() == target_personnel:
                    matching_columns.append(col_idx)
            
            if not matching_columns:
                st.error(f"❌ 未找到人事號: {target_personnel}")
                return
            
            # 計算加班時數
            weekday_hours, weekend_hours, total_hours, records, daily_breakdown = calculate_overtime_summary(
                target_personnel, year, month_value, matching_columns
            )
            
            # 儲存查詢結果
            st.session_state.last_query_result = {
                'target_personnel': target_personnel,
                'year': year,
                'month': month_value,
                'matching_columns': matching_columns,
                'daily_breakdown': daily_breakdown,
                'weekday_hours': weekday_hours,
                'weekend_hours': weekend_hours,
                'total_hours': total_hours
            }
        
        # 顯示結果
        st.success("✅ 查詢完成！")
        
        # 顯示版本資訊
        st.info(f"📋 使用版本: {st.session_state.version_info['version_name']}")
        
        # 顯示自定義假日資訊
        if st.session_state.custom_holidays:
            current_month_holidays = {k: v for k, v in st.session_state.custom_holidays.items()
                                    if k.startswith(f"{year}-{month_value:02d}-")}
            if current_month_holidays:
                st.info(f"🏖️ 本月自定義假日 ({len(current_month_holidays)} 天)")
                holiday_list = []
                for date_key, desc in sorted(current_month_holidays.items()):
                    holiday_list.append(f"• {date_key}: {desc}")
                st.markdown("\n".join(holiday_list))
        
        # 統計結果卡片
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("平日加班時數", f"{weekday_hours:.1f} 小時")
        with col2:
            st.metric("假日加班時數", f"{weekend_hours:.1f} 小時")
        with col3:
            st.metric("總加班時數", f"{total_hours:.1f} 小時")
        
        # 詳細每日資料
        if daily_breakdown:
            st.subheader("📅 詳細每日加班記錄")
            
            # 創建表格數據
            table_data = []
            for date_str, hours in sorted(daily_breakdown.items()):
                if hours > 0:
                    date_parts = date_str.split('/')
                    check_year = int(date_parts[0])
                    check_month = int(date_parts[1])
                    check_day = int(date_parts[2])
                    day_type, is_weekend = get_day_type(check_year, check_month, check_day)
                    
                    table_data.append({
                        '日期': date_str,
                        '星期': day_type,
                        '加班時數': f"{hours:.1f}小時",
                        '類型': '假日' if is_weekend else '平日'
                    })
            
            if table_data:
                df_display = pd.DataFrame(table_data)
                st.dataframe(df_display, use_container_width=True)
    
    # Excel 匯出按鈕 - 移到查詢結果外面，但需要有查詢結果才能使用
    if st.session_state.last_query_result is not None:
        st.subheader("📊 匯出報表")
        
        # 創建下載按鈕的容器
        export_container = st.container()
        
        with export_container:
            if st.button("📊 產生並下載Excel報表", type="secondary", key="export_excel_btn"):
                with st.spinner("📊 正在產生Excel報表..."):
                    result = st.session_state.last_query_result
                    success, file_content_or_error, weekday_total, weekend_total, total_hours_export, row_count = export_to_excel(
                        result['target_personnel'],
                        result['year'],
                        result['month'],
                        result['matching_columns'],
                        result['daily_breakdown']
                    )
                    
                    if success:
                        filename = f"{result['target_personnel']}_{result['year']}年{result['month']:02d}月_加班時數統計.xlsx"
                        
                        st.success("✅ Excel報表產生成功！")
                        st.info(f"📋 使用版本: {st.session_state.version_info['version_name']}")
                        st.info(f"📋 包含資料: {row_count} 天")
                        
                        # 提供下載按鈕
                        st.download_button(
                            label="📥 下載Excel檔案",
                            data=file_content_or_error.getvalue(),
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_excel_btn"
                        )
                    else:
                        st.error(f"❌ Excel匯出失敗: {file_content_or_error}")

def holiday_management_page():
    """自定義假日管理頁面"""
    st.header("🗓️ 自定義假日管理")
    
    # 新增假日區域
    st.subheader("➕ 新增自定義假日")
    
    with st.form("add_holiday_form"):
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            holiday_year = st.number_input("年份", min_value=2020, max_value=2030, 
                                         value=datetime.now().year)
        
        with col2:
            holiday_month = st.selectbox("月份", 
                                       [(i, f"{i}月") for i in range(1, 13)], 
                                       index=datetime.now().month-1,
                                       format_func=lambda x: x[1])
        
        with col3:
            holiday_day = st.number_input("日期", min_value=1, max_value=31, value=1)
        
        with col4:
            holiday_reason = st.text_input("假日原因", value="自定義假日")
        
        col_add, col_remove = st.columns(2)
        with col_add:
            add_holiday = st.form_submit_button("➕ 新增假日", type="primary")
        with col_remove:
            remove_holiday = st.form_submit_button("❌ 移除假日", type="secondary")
    
    # 處理新增假日
    if add_holiday:
        year_val = holiday_year
        month_val = holiday_month[0]
        day_val = holiday_day
        reason = holiday_reason.strip() if holiday_reason.strip() else "自定義假日"
        
        try:
            # 驗證日期
            test_date = date(year_val, month_val, day_val)
            date_key = f"{year_val}-{month_val:02d}-{day_val:02d}"
            
            weekdays = ['一', '二', '三', '四', '五', '六', '日']
            weekday = weekdays[test_date.weekday()]
            
            st.session_state.custom_holidays[date_key] = f"{reason}({weekday})"
            st.success(f"✅ 已新增自定義假日: {date_key} {reason}({weekday})")
            
        except ValueError:
            st.error(f"❌ 無效日期: {year_val}-{month_val:02d}-{day_val:02d}")
    
    # 處理移除假日
    if remove_holiday:
        year_val = holiday_year
        month_val = holiday_month[0]
        day_val = holiday_day
        date_key = f"{year_val}-{month_val:02d}-{day_val:02d}"
        
        if date_key in st.session_state.custom_holidays:
            removed = st.session_state.custom_holidays.pop(date_key)
            st.success(f"✅ 已移除自定義假日: {date_key} ({removed})")
        else:
            st.warning(f"⚠️ 該日期不是自定義假日: {date_key}")
    
    # 管理現有假日
    st.subheader("📅 目前設定的自定義假日")
    
    col_clear, col_refresh = st.columns(2)
    with col_clear:
        if st.button("🗑️ 清除所有假日", type="secondary"):
            if st.session_state.custom_holidays:
                st.session_state.custom_holidays.clear()
                st.success("✅ 已清除所有自定義假日")
            else:
                st.info("📅 目前沒有設定任何自定義假日")
    
    # 顯示假日清單
    if st.session_state.custom_holidays:
        st.write(f"**目前共有 {len(st.session_state.custom_holidays)} 個自定義假日:**")
        
        # 轉換為表格顯示
        holiday_data = []
        for date_key, description in sorted(st.session_state.custom_holidays.items()):
            holiday_data.append({
                '日期': date_key,
                '描述': description,
                '年': date_key.split('-')[0],
                '月': date_key.split('-')[1],
                '日': date_key.split('-')[2]
            })
        
        df_holidays = pd.DataFrame(holiday_data)
        st.dataframe(df_holidays[['日期', '描述']], use_container_width=True)
        
        # 按月份分組顯示
        st.subheader("📊 按月份分組")
        months_dict = {}
        for date_key, description in st.session_state.custom_holidays.items():
            year_month = date_key[:7]  # YYYY-MM
            if year_month not in months_dict:
                months_dict[year_month] = []
            months_dict[year_month].append(f"{date_key}: {description}")
        
        for year_month, holidays in sorted(months_dict.items()):
            with st.expander(f"📅 {year_month} ({len(holidays)} 天)"):
                for holiday in holidays:
                    st.write(f"• {holiday}")
    else:
        st.info("📅 目前沒有設定任何自定義假日")

if __name__ == "__main__":
    main()