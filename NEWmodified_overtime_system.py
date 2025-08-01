# äº’å‹•å¼å“¡å·¥ç­è¡¨åŠ ç­æ™‚æ•¸çµ±è¨ˆç³»çµ± (Streamlitç‰ˆ) - ä¿®æ”¹ç‰ˆ
import pandas as pd
from datetime import datetime, date, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import streamlit as st
import warnings
import calendar
import io
import base64
warnings.filterwarnings('ignore')

# ===== Streamlit é é¢é…ç½® =====
st.set_page_config(
    page_title="å“¡å·¥ç­è¡¨åŠ ç­æ™‚æ•¸çµ±è¨ˆç³»çµ±",
    page_icon="ğŸ¢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ===== åˆå§‹åŒ– Session State =====
if 'df' not in st.session_state:
    st.session_state.df = None
if 'shift_dict' not in st.session_state:
    st.session_state.shift_dict = {}
if 'custom_holidays' not in st.session_state:
    st.session_state.custom_holidays = {}
if 'last_query_result' not in st.session_state:
    st.session_state.last_query_result = None
if 'current_page' not in st.session_state:
    st.session_state.current_page = "è¼‰å…¥ç­è¡¨è³‡æ–™"
if 'preview_data' not in st.session_state:
    st.session_state.preview_data = None

# ===== å·¥å…·å‡½æ•¸ =====
def convert_google_sheet_url(url):
    """å°‡ Google Sheets URL è½‰æ›ç‚ºå¯ç›´æ¥è®€å–çš„ CSV URL"""
    if '/d/' in url:
        sheet_id = url.split('/d/')[1].split('/')[0]
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    return None

@st.cache_data
def load_data_from_urls(main_sheet_url):
    """ä½¿ç”¨è‡ªå®šç¾©URLè¼‰å…¥è³‡æ–™"""
    # ç­ç¨®å°ç…§è¡¨å›ºå®šä½¿ç”¨é è¨­é€£çµ
    default_shift_sheet_url = "https://docs.google.com/spreadsheets/d/1JfhGZYRBWj6yp90o-sA0DrhzkcEM1Wfd_vqiEZEYd5c/edit?usp=sharing"

    main_csv_url = convert_google_sheet_url(main_sheet_url)
    shift_csv_url = convert_google_sheet_url(default_shift_sheet_url)
    
    if not main_csv_url or not shift_csv_url:
        return None, None, "âŒ ç„¡æ•ˆçš„Google Sheets URLæ ¼å¼"

    try:
        # è®€å–å“¡å·¥ç­è¡¨
        df_full = pd.read_csv(main_csv_url)
        df = df_full.iloc[:36, :83]  # é¸å– A1:CE36 ç¯„åœ

        # è®€å–ç­ç¨®å°ç…§è¡¨
        shift_df = pd.read_csv(shift_csv_url)

        # å»ºç«‹ç­ç¨®å­—å…¸
        shift_dict = {}
        for index, row in shift_df.iterrows():
            shift_type = str(row.iloc[0]).strip()
            overtime_hours_1 = row.iloc[1]
            overtime_hours_2 = row.iloc[2]
            cross_day_hours = row.iloc[3] if len(row) > 3 else None

            shift_dict[shift_type] = {
                'overtime_hours_1': overtime_hours_1,
                'overtime_hours_2': overtime_hours_2,
                'cross_day_hours': cross_day_hours
            }

        return df, shift_dict, f"âœ… è³‡æ–™è®€å–æˆåŠŸï¼ç­è¡¨: {df.shape}, ç­ç¨®: {len(shift_dict)} ç¨®"
    except Exception as e:
        return None, None, f"âŒ è³‡æ–™è®€å–å¤±æ•—: {e}"

def calculate_hours(time_range):
    """è¨ˆç®—æ™‚é–“ç¯„åœçš„å°æ™‚æ•¸"""
    try:
        if not time_range or pd.isna(time_range):
            return None

        time_str = str(time_range).strip()

        # è™•ç†é€—è™Ÿä½œç‚ºå°æ•¸é»çš„æƒ…æ³
        if ',' in time_str and '-' not in time_str:
            try:
                hours = float(time_str.replace(',', '.'))
                return hours
            except:
                pass

        # å¦‚æœä¸åŒ…å«åˆ†éš”ç¬¦ï¼Œå¯èƒ½æ˜¯å–®ç´”çš„å°æ™‚æ•¸
        if '-' not in time_str:
            try:
                hours = float(time_str)
                return hours
            except:
                return None

        # æ¸…ç†æ™‚é–“å­—ä¸²
        time_str = time_str.replace(' ', '').replace(',', '')

        if '-' in time_str:
            parts = time_str.split('-')
            if len(parts) != 2:
                return None
            start_str, end_str = parts
        else:
            return None

        def parse_time_hhmm(time_str):
            """è§£æ HH:MM æ ¼å¼çš„æ™‚é–“"""
            time_str = time_str.strip()

            # è™•ç† HH:MM æ ¼å¼
            if ':' in time_str:
                try:
                    parts = time_str.split(':')
                    if len(parts) == 2:
                        hour = int(parts[0])
                        minute = int(parts[1])
                        if 0 <= hour <= 23 and 0 <= minute <= 59:
                            return hour, minute
                except ValueError:
                    pass

            # è™•ç† HHMM æ ¼å¼
            if len(time_str) == 4 and time_str.isdigit():
                try:
                    hour = int(time_str[:2])
                    minute = int(time_str[2:])
                    if 0 <= hour <= 23 and 0 <= minute <= 59:
                        return hour, minute
                except ValueError:
                    pass

            # è™•ç† HH æ ¼å¼
            if time_str.isdigit() and 1 <= len(time_str) <= 2:
                try:
                    hour = int(time_str)
                    if 0 <= hour <= 23:
                        return hour, 0
                except ValueError:
                    pass

            # è™•ç†å¸¶å°æ•¸é»çš„å°æ™‚æ•¸æ ¼å¼
            try:
                hour_decimal = float(time_str)
                if 0 <= hour_decimal <= 24:
                    hour = int(hour_decimal)
                    minute = int((hour_decimal - hour) * 60)
                    if 0 <= hour <= 23 and 0 <= minute <= 59:
                        return hour, minute
            except ValueError:
                pass

            return None, None

        # è§£æé–‹å§‹å’ŒçµæŸæ™‚é–“
        start_hour, start_min = parse_time_hhmm(start_str)
        end_hour, end_min = parse_time_hhmm(end_str)

        if start_hour is None or end_hour is None:
            return None

        # è½‰æ›ç‚ºåˆ†é˜
        start_minutes = start_hour * 60 + start_min
        end_minutes = end_hour * 60 + end_min

        # è™•ç†è·¨æ—¥æƒ…æ³
        if end_minutes <= start_minutes:
            end_minutes += 24 * 60

        # è¨ˆç®—æ™‚æ•¸
        total_minutes = end_minutes - start_minutes
        hours = total_minutes / 60

        return hours

    except Exception as e:
        return None

def get_day_type(year, month, day):
    """åˆ¤æ–·æ—¥æœŸæ˜¯å¹³æ—¥é‚„æ˜¯å‡æ—¥ï¼ˆå«è‡ªå®šç¾©å‡æ—¥ï¼‰"""
    try:
        # æª¢æŸ¥æ˜¯å¦ç‚ºè‡ªå®šç¾©å‡æ—¥
        date_key = f"{year}-{month:02d}-{day:02d}"
        if date_key in st.session_state.custom_holidays:
            return st.session_state.custom_holidays[date_key], True

        # ä¸€èˆ¬é€±æœ«åˆ¤æ–·
        current_date = date(year, month, day)
        weekday = current_date.weekday()  # 0=æ˜ŸæœŸä¸€, 6=æ˜ŸæœŸæ—¥

        if weekday == 5:  # æ˜ŸæœŸå…­
            return "å‡æ—¥(å…­)", True
        elif weekday == 6:  # æ˜ŸæœŸæ—¥
            return "å‡æ—¥(æ—¥)", True
        else:  # æ˜ŸæœŸä¸€åˆ°æ˜ŸæœŸäº”
            weekdays = ["ä¸€", "äºŒ", "ä¸‰", "å››", "äº”"]
            return f"å¹³æ—¥({weekdays[weekday]})", False
    except ValueError:
        return "ç„¡æ•ˆæ—¥æœŸ", False

def extract_chinese_note(time_string):
    """å¾æ™‚é–“å­—ä¸²ä¸­æå–ä¸­æ–‡è¨»è¨˜"""
    import re

    if not time_string:
        return "è‡¨åºŠæ¥­å‹™"

    chinese_pattern = r'[\u4e00-\u9fff]+|\([^\)]*[\u4e00-\u9fff][^\)]*\)'
    chinese_matches = re.findall(chinese_pattern, time_string)

    if chinese_matches:
        chinese_note = chinese_matches[0]
        chinese_note = chinese_note.replace('(', '').replace(')', '')
        return chinese_note
    else:
        return "è‡¨åºŠæ¥­å‹™"

def generate_schedule_preview(target_personnel, year, month, matching_columns):
    """ç”Ÿæˆç­è¡¨é è¦½è³‡æ–™"""
    df = st.session_state.df
    shift_dict = st.session_state.shift_dict
    
    preview_data = []
    
    for day in range(1, 32):
        try:
            current_date = date(year, month, day)
            date_str = f"{year}/{month:02d}/{day:02d}"
            day_type, is_weekend = get_day_type(year, month, day)
            
            # æ”¶é›†æ‰€æœ‰åŒ¹é…æ¬„ä½çš„ç­æ¬¡è³‡æ–™
            shifts = []
            for col_idx in matching_columns:
                column_data = df.iloc[:, col_idx]
                row_idx = day + 2
                
                if row_idx < len(column_data):
                    value = column_data.iloc[row_idx]
                    shift_value = str(value).strip()
                    
                    if shift_value and pd.notna(value) and shift_value != 'nan':
                        shifts.append(shift_value)
            
            shift_display = ', '.join(shifts) if shifts else 'ä¼‘å‡'
            
            preview_data.append({
                'æ—¥æœŸ': f"{day:02d}",
                'æ˜ŸæœŸ': day_type,
                'ç­æ¬¡': shift_display,
                'é¡å‹': 'å‡æ—¥' if is_weekend else 'å¹³æ—¥'
            })
            
        except ValueError:
            continue
    
    return preview_data

def calculate_overtime_summary(target_personnel, year, month, matching_columns):
    """è¨ˆç®—æŒ‡å®šäººå“¡çš„åŠ ç­æ™‚æ•¸çµ±è¨ˆ"""
    df = st.session_state.df
    shift_dict = st.session_state.shift_dict
    
    weekday_overtime = 0.0
    weekend_overtime = 0.0
    total_overtime = 0.0
    daily_records = []
    cross_day_records = defaultdict(float)
    worked_weekdays = set()

    # å°æ¯å€‹åŒ¹é…çš„æ¬„ä½é€²è¡Œè™•ç†
    for col_idx in matching_columns:
        column_data = df.iloc[:, col_idx]

        # è™•ç†æ¯ä¸€å¤©
        for day in range(1, 32):
            try:
                current_date = date(year, month, day)
            except ValueError:
                continue

            row_idx = day + 2

            if row_idx < len(column_data):
                value = column_data.iloc[row_idx]
                date_str = f"{year}/{month:02d}/{day:02d}"
                day_type, is_weekend = get_day_type(year, month, day)

                shift_value = str(value).strip()

                # è¨˜éŒ„æœ‰ä¸Šç­çš„å¹³æ—¥
                if shift_value and pd.notna(value) and shift_value != 'nan' and not is_weekend:
                    worked_weekdays.add(date_str)

                if shift_value in shift_dict and pd.notna(value) and shift_value:
                    shift_info = shift_dict[shift_value]
                    overtime_1 = shift_info['overtime_hours_1']
                    overtime_2 = shift_info['overtime_hours_2']
                    cross_day = shift_info['cross_day_hours']

                    current_day_overtime = 0.0
                    next_day_overtime = 0.0

                    # è¨ˆç®—ç•¶å¤©åŠ ç­æ™‚æ•¸
                    if pd.notna(overtime_1) and str(overtime_1).strip():
                        calculated_hours_1 = calculate_hours(str(overtime_1))
                        if calculated_hours_1 is not None:
                            current_day_overtime += calculated_hours_1

                    if pd.notna(overtime_2) and str(overtime_2).strip():
                        calculated_hours_2 = calculate_hours(str(overtime_2))
                        if calculated_hours_2 is not None:
                            current_day_overtime += calculated_hours_2

                    # è¨ˆç®—è·¨å¤©æ™‚æ•¸
                    if pd.notna(cross_day) and str(cross_day).strip():
                        cross_day_hours = calculate_hours(str(cross_day))
                        if cross_day_hours is not None:
                            next_day_overtime = cross_day_hours
                            next_date = current_date + timedelta(days=1)
                            next_date_str = f"{next_date.year}/{next_date.month:02d}/{next_date.day:02d}"
                            cross_day_records[next_date_str] += next_day_overtime

                    # è¨˜éŒ„ç•¶å¤©è³‡æ–™
                    if current_day_overtime > 0 or next_day_overtime > 0:
                        daily_records.append({
                            'date': date_str,
                            'day_type': day_type,
                            'is_weekend': is_weekend,
                            'shift': shift_value,
                            'current_day_overtime': current_day_overtime,
                            'cross_day_overtime': next_day_overtime
                        })

    # å»ºç«‹å®Œæ•´çš„æ—¥æœŸåŠ ç­æ™‚æ•¸è¨˜éŒ„
    final_daily_overtime = defaultdict(float)

    # å…ˆåŠ å…¥ç•¶å¤©çš„åŠ ç­æ™‚æ•¸
    for record in daily_records:
        date_str = record['date']
        current_overtime = record['current_day_overtime']
        if current_overtime > 0:
            final_daily_overtime[date_str] += current_overtime

    # å†åŠ å…¥è·¨å¤©æ™‚æ•¸
    for date_str, cross_hours in cross_day_records.items():
        final_daily_overtime[date_str] += cross_hours

    # è¨ˆç®—å¹³æ—¥å’Œå‡æ—¥ç¸½æ™‚æ•¸
    weekday_dates = []
    weekend_dates = []
    
    for date_str, total_hours in final_daily_overtime.items():
        try:
            date_parts = date_str.split('/')
            check_year = int(date_parts[0])
            check_month = int(date_parts[1])
            check_day = int(date_parts[2])

            day_type, is_weekend = get_day_type(check_year, check_month, check_day)

            if is_weekend:
                weekend_overtime += total_hours
                weekend_dates.append((date_str, total_hours))
            else:
                weekday_overtime += total_hours
                weekday_dates.append((date_str, total_hours))

            total_overtime += total_hours

        except (ValueError, IndexError):
            continue

    # å¦‚æœå¹³æ—¥ç¸½æ™‚æ•¸è¶…é46å°æ™‚ï¼Œåˆªé™¤æ—¥æœŸè®“ç¸½æ™‚æ•¸ç­‰æ–¼æˆ–æ¥è¿‘46å°æ™‚
    if weekday_overtime > 46:
        # æŒ‰æ™‚æ•¸æ’åºï¼ˆå¾å°åˆ°å¤§ï¼‰ï¼Œå„ªå…ˆåˆªé™¤è¼ƒå°çš„æ™‚æ•¸
        weekday_dates.sort(key=lambda x: x[1])
        
        excess_hours = weekday_overtime - 46
        removed_hours = 0.0
        
        for date_str, hours in weekday_dates:
            if removed_hours + hours <= excess_hours:
                # å®Œå…¨ç§»é™¤é€™ä¸€å¤©
                final_daily_overtime[date_str] = 0.0
                removed_hours += hours
                weekday_overtime -= hours
                total_overtime -= hours
                
                if removed_hours >= excess_hours:
                    break
            elif removed_hours < excess_hours:
                # éƒ¨åˆ†ç§»é™¤
                remaining_to_remove = excess_hours - removed_hours
                final_daily_overtime[date_str] -= remaining_to_remove
                weekday_overtime -= remaining_to_remove
                total_overtime -= remaining_to_remove
                break

    # æª¢æŸ¥å¹³æ—¥åŠ ç­æ™‚æ•¸æ˜¯å¦å°æ–¼46å°æ™‚ï¼Œå¦‚æœæ˜¯å‰‡è‡ªå‹•è£œè¶³
    elif weekday_overtime < 46:
        shortage = 46 - weekday_overtime

        # æ‰¾å‡ºæ‰€æœ‰è©²æœˆçš„å¹³æ—¥ä¸”æ²’æœ‰ä¸Šç­çš„æ—¥æœŸ
        available_weekdays = []
        for day in range(1, 32):
            try:
                check_date = date(year, month, day)
                date_str = f"{year}/{month:02d}/{day:02d}"
                day_type, is_weekend = get_day_type(year, month, day)
                weekday_num = check_date.weekday()

                if not is_weekend and date_str not in worked_weekdays:
                    if weekday_num in [1, 3]:  # é€±äºŒã€é€±å››
                        priority = 1
                    elif weekday_num in [0, 2, 4]:  # é€±ä¸€ã€é€±ä¸‰ã€é€±äº”
                        priority = 2
                    else:
                        priority = 3

                    available_weekdays.append((date_str, day_type, weekday_num, priority))
            except ValueError:
                continue

        # æŒ‰å„ªå…ˆé †åºæ’åº
        available_weekdays.sort(key=lambda x: (x[3], x[0]))

        if available_weekdays:
            days_needed = int(shortage / 2) + (1 if shortage % 2 > 0 else 0)

            for i, (date_str, day_type, weekday_num, priority) in enumerate(available_weekdays):
                if i < days_needed:
                    final_daily_overtime[date_str] += 2.0
                    weekday_overtime += 2.0
                    total_overtime += 2.0

    return weekday_overtime, weekend_overtime, total_overtime, daily_records, final_daily_overtime

def export_to_excel(target_personnel, year, month, matching_columns, export_data):
    """å°å‡ºExcelå ±è¡¨"""
    try:
        df = st.session_state.df
        shift_dict = st.session_state.shift_dict
        
        excel_data = []
        date_time_strings = defaultdict(list)

        # æ”¶é›†åŸå§‹æ™‚é–“å­—ä¸²
        for col_idx in matching_columns:
            column_data = df.iloc[:, col_idx]

            for day in range(1, 32):
                try:
                    current_date = date(year, month, day)
                    date_str = f"{year}/{month:02d}/{day:02d}"
                    row_idx = day + 2

                    if row_idx < len(column_data):
                        value = column_data.iloc[row_idx]
                        shift_value = str(value).strip()

                        if shift_value in shift_dict and pd.notna(value) and shift_value:
                            shift_info = shift_dict[shift_value]

                            current_day_time_strings = []

                            if pd.notna(shift_info['overtime_hours_1']) and str(shift_info['overtime_hours_1']).strip():
                                current_day_time_strings.append(str(shift_info['overtime_hours_1']).strip())

                            if pd.notna(shift_info['overtime_hours_2']) and str(shift_info['overtime_hours_2']).strip():
                                current_day_time_strings.append(str(shift_info['overtime_hours_2']).strip())

                            if current_day_time_strings:
                                date_time_strings[date_str].extend(current_day_time_strings)

                            if pd.notna(shift_info['cross_day_hours']) and str(shift_info['cross_day_hours']).strip():
                                cross_day_time_str = str(shift_info['cross_day_hours']).strip()
                                next_date = current_date + timedelta(days=1)
                                next_date_str = f"{next_date.year}/{next_date.month:02d}/{next_date.day:02d}"
                                date_time_strings[next_date_str].append(cross_day_time_str)

                except ValueError:
                    continue

        # å»ºç«‹Excelè³‡æ–™
        for day in range(1, 32):
            try:
                current_date = date(year, month, day)
                date_str = f"{year}/{month:02d}/{day:02d}"
                day_type, is_weekend = get_day_type(year, month, day)

                time_strings = date_time_strings.get(date_str, [])
                original_time_str = ",".join(time_strings) if time_strings else ""

                weekday_hours = 0.0
                weekend_hours = 0.0

                if date_str in export_data:
                    total_hours = export_data[date_str]
                    if is_weekend:
                        weekend_hours = total_hours
                        
                        # ä¿®æ”¹å¾Œçš„å‡æ—¥åŠ ç­é‚è¼¯
                        if weekend_hours <= 3 and weekend_hours > 0:
                            if original_time_str:
                                # æå–ç¬¬ä¸€å€‹æ™‚é–“çš„çµæŸæ™‚é–“
                                first_time_part = original_time_str.split(',')[0]
                                if '-' in first_time_part:
                                    end_time = first_time_part.split('-')[1].strip()
                                    try:
                                        # è§£æçµæŸæ™‚é–“
                                        if ':' in end_time:
                                            end_hour = int(end_time.split(':')[0])
                                            end_minute = int(end_time.split(':')[1])
                                        else:
                                            end_hour = int(end_time[:2]) if len(end_time) >= 2 else int(end_time)
                                            end_minute = 0
                                        
                                        # å¦‚æœçµæŸæ™‚é–“åœ¨05:00ä¹‹å‰ï¼Œåœ¨å¾Œé¢åŠ 2å°æ™‚
                                        if end_hour < 5:
                                            new_start_hour = end_hour
                                            new_start_minute = end_minute
                                            new_end_hour = end_hour + 2
                                            if new_end_hour >= 24:
                                                new_end_hour -= 24
                                            
                                            new_time_part = f"{new_start_hour:02d}:{new_start_minute:02d}-{new_end_hour:02d}:{end_minute:02d}(æ’°å¯«ç—…æ­·)"
                                            original_time_str = original_time_str + "," + new_time_part
                                        else:
                                            # å¦‚æœçµæŸæ™‚é–“åœ¨05:00ä¹‹å¾Œï¼Œåœ¨å‰é¢åŠ 2å°æ™‚
                                            start_time = first_time_part.split('-')[0].strip()
                                            if ':' in start_time:
                                                start_hour = int(start_time.split(':')[0])
                                                start_minute = int(start_time.split(':')[1])
                                            else:
                                                start_hour = int(start_time[:2]) if len(start_time) >= 2 else int(start_time)
                                                start_minute = 0
                                            
                                            new_start_hour = start_hour - 2
                                            if new_start_hour < 0:
                                                new_start_hour += 24
                                            
                                            new_time_part = f"{new_start_hour:02d}:{start_minute:02d}-{start_hour:02d}:{start_minute:02d}(æ’°å¯«ç—…æ­·)"
                                            original_time_str = new_time_part + "," + original_time_str
                                        
                                        weekend_hours += 2.0
                                        
                                    except (ValueError, IndexError):
                                        # å¦‚æœè§£æå¤±æ•—ï¼Œä½¿ç”¨é è¨­
                                        original_time_str = "12:00-14:00(æ’°å¯«ç—…æ­·)," + original_time_str
                                        weekend_hours += 2.0
                                else:
                                    original_time_str = "12:00-14:00(æ’°å¯«ç—…æ­·)," + original_time_str
                                    weekend_hours += 2.0
                            else:
                                original_time_str = "12:00-14:00(æ’°å¯«ç—…æ­·)"
                                weekend_hours += 2.0
                    else:
                        weekday_hours = total_hours

                work_type = ""
                if date_str in export_data and not original_time_str:
                    original_time_str = "14:00-16:00(æœƒè­°)"
                    work_type = "æœƒè­°"
                else:
                    work_type = extract_chinese_note(original_time_str)

                if original_time_str or weekday_hours > 0 or weekend_hours > 0:
                    excel_data.append({
                        'æ—¥æœŸ': f"{day:02d}",
                        'åŸå§‹æ™‚é–“å­—ä¸²': original_time_str,
                        'å¹³æ—¥æ™‚æ•¸': weekday_hours,
                        'å‡æ—¥æ™‚æ•¸': weekend_hours,
                        'å·¥ä½œé¡å‹': work_type
                    })

            except ValueError:
                continue

        # å»ºç«‹Excelæª”æ¡ˆ
        df_excel = pd.DataFrame(excel_data)

        # å‰µå»ºExcelå…§å®¹åˆ°å…§å­˜
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{target_personnel}åŠ ç­çµ±è¨ˆ"

        # è¨­å®šæ¨™é¡Œ
        headers = ['æ—¥æœŸ', 'åŸå§‹æ™‚é–“å­—ä¸²', 'å¹³æ—¥æ™‚æ•¸', 'å‡æ—¥æ™‚æ•¸', 'å·¥ä½œé¡å‹']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', size=12)

        # è¨­å®šé‚Šæ¡†
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # å¡«å…¥è³‡æ–™
        for row_idx, row_data in enumerate(df_excel.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                if col_idx in [3, 4]:  # å¹³æ—¥æ™‚æ•¸ã€å‡æ—¥æ™‚æ•¸
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if value > 0:
                        cell.number_format = '0.0'
                elif col_idx == 5:  # å·¥ä½œé¡å‹
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # èª¿æ•´æ¬„å¯¬
        column_widths = [8, 30, 12, 12, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # æ·»åŠ çµ±è¨ˆ
        total_weekday = df_excel['å¹³æ—¥æ™‚æ•¸'].sum()
        total_weekend = df_excel['å‡æ—¥æ™‚æ•¸'].sum()
        total_hours = total_weekday + total_weekend

        last_row = len(df_excel) + 3

        ws.cell(row=last_row, column=1, value="çµ±è¨ˆç¸½è¨ˆ").font = Font(bold=True, size=12)
        ws.cell(row=last_row, column=1).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        ws.cell(row=last_row + 1, column=1, value="å¹³æ—¥åŠ ç­ç¸½æ™‚æ•¸:")
        ws.cell(row=last_row + 1, column=2, value=f"{total_weekday:.1f} å°æ™‚")

        ws.cell(row=last_row + 2, column=1, value="å‡æ—¥åŠ ç­ç¸½æ™‚æ•¸:")
        ws.cell(row=last_row + 2, column=2, value=f"{total_weekend:.1f} å°æ™‚")

        ws.cell(row=last_row + 3, column=1, value="ç¸½åŠ ç­æ™‚æ•¸:")
        ws.cell(row=last_row + 3, column=2, value=f"{total_hours:.1f} å°æ™‚")
        ws.cell(row=last_row + 3, column=2).font = Font(bold=True)

        wb.save(output)
        output.seek(0)

        return True, output, total_weekday, total_weekend, total_hours, len(df_excel)

    except Exception as e:
        return False, str(e), 0, 0, 0, 0

# ===== ä¸»è¦ç•Œé¢ =====
def main():
    st.title("ğŸ¢ å“¡å·¥ç­è¡¨åŠ ç­æ™‚æ•¸çµ±è¨ˆç³»çµ±")
    
    # å´é‚Šæ¬„
    with st.sidebar:
        st.header("ğŸ“‹ ç³»çµ±åŠŸèƒ½")
        
        if st.button("ğŸ“¥ è¼‰å…¥ç­è¡¨è³‡æ–™", type="primary" if st.session_state.current_page == "è¼‰å…¥ç­è¡¨è³‡æ–™" else "secondary"):
            st.session_state.current_page = "è¼‰å…¥ç­è¡¨è³‡æ–™"
            st.rerun()
        
        if st.button("ğŸ” æŸ¥è©¢åŠ ç­æ™‚æ•¸", type="primary" if st.session_state.current_page == "æŸ¥è©¢åŠ ç­æ™‚æ•¸" else "secondary"):
            st.session_state.current_page = "æŸ¥è©¢åŠ ç­æ™‚æ•¸"
            st.rerun()
        
        if st.button("ğŸ—“ï¸ è‡ªå®šç¾©å‡æ—¥ç®¡ç†", type="primary" if st.session_state.current_page == "è‡ªå®šç¾©å‡æ—¥ç®¡ç†" else "secondary"):
            st.session_state.current_page = "è‡ªå®šç¾©å‡æ—¥ç®¡ç†"
            st.rerun()
        
        # æ–°å¢é–‹å•Ÿç©ºç™½åŠ ç­å–®æŒ‰éˆ•
        st.markdown("---")
        st.markdown("### ğŸ“ å…¶ä»–åŠŸèƒ½")
        
        # ä½¿ç”¨ st.link_button æˆ–è€…é¡¯ç¤ºé€£çµ
        overtime_form_url = "https://docs.google.com/document/d/1T75rw_3hQtIaBTGMFxa09G93Atihf4h-883Kg1tqPpo/edit?usp=sharing"
        
        # æ–¹æ³•1: ä½¿ç”¨ markdown å‰µå»ºå¯é»æ“Šçš„é€£çµ
        st.markdown(f"""
        <a href="{overtime_form_url}" target="_blank">
            <button style="
                background-color: #f0f2f6;
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 8px 16px;
                font-size: 14px;
                cursor: pointer;
                width: 100%;
                text-align: center;
                color: #262730;
            ">
                ğŸ“„ é–‹å•Ÿç©ºç™½åŠ ç­å–®
            </button>
        </a>
        """, unsafe_allow_html=True)
        
        # æ–¹æ³•2: ä¹Ÿæä¾›ç›´æ¥çš„æ–‡å­—é€£çµä½œç‚ºå‚™ç”¨
        st.markdown(f"ğŸ”— [é»æ­¤é–‹å•Ÿç©ºç™½åŠ ç­å–®]({overtime_form_url})", help="åœ¨æ–°åˆ†é é–‹å•ŸGoogleæ–‡ä»¶åŠ ç­å–®")
    
    # é¡¯ç¤ºè¼‰å…¥ç‹€æ…‹
    if st.session_state.df is not None:
        personnel_numbers = st.session_state.df.iloc[1, :].tolist()
        personnel_count = sum(1 for num in personnel_numbers if pd.notna(num) and str(num).strip())
        st.success(f"ğŸ“‹ ç­è¡¨å·²è¼‰å…¥ - å…± {personnel_count} å€‹äººäº‹è™Ÿï¼Œ{len(st.session_state.shift_dict)} ç¨®ç­æ¬¡")
    else:
        st.warning("ğŸ“‹ å°šæœªè¼‰å…¥ä»»ä½•ç­è¡¨è³‡æ–™")
    
    # æ ¹æ“šç•¶å‰é é¢é¡¯ç¤ºå°æ‡‰å…§å®¹
    if st.session_state.current_page == "è¼‰å…¥ç­è¡¨è³‡æ–™":
        load_data_page()
    elif st.session_state.current_page == "æŸ¥è©¢åŠ ç­æ™‚æ•¸":
        query_page()
    elif st.session_state.current_page == "è‡ªå®šç¾©å‡æ—¥ç®¡ç†":
        holiday_management_page()

def load_data_page():
    """è¼‰å…¥è³‡æ–™é é¢"""
    st.header("ğŸ“¥ è¼‰å…¥é›²ç«¯ç­è¡¨è³‡æ–™")
    
    with st.form("load_data_form"):
        main_sheet_url = st.text_area(
            "å“¡å·¥ç­è¡¨ Google Sheets é€£çµ",
            placeholder="è«‹è²¼ä¸Šå“¡å·¥ç­è¡¨çš„ Google Sheets å®Œæ•´é€£çµ",
            help="è«‹ç¢ºä¿ Google Sheets å·²è¨­å®šç‚ºã€ŒçŸ¥é“é€£çµçš„ä½¿ç”¨è€…ã€å¯æª¢è¦–"
        )
        
        st.info("ğŸ“‹ ç­ç¨®å°ç…§è¡¨: ç³»çµ±å°‡è‡ªå‹•ä½¿ç”¨é è¨­çš„ç­ç¨®å°ç…§è¡¨")
        
        col1, col2 = st.columns(2)
        with col1:
            submit_button = st.form_submit_button("ğŸ“¥ è¼‰å…¥é›²ç«¯ç­è¡¨", type="primary")
        with col2:
            load_default = st.form_submit_button("ğŸ”„ è¼‰å…¥é è¨­ç­è¡¨")
    
    # è™•ç†è¼‰å…¥é è¨­ç­è¡¨
    if load_default:
        main_sheet_url = "https://docs.google.com/spreadsheets/d/1U8qLraVCRKJhySk0y93I_POP_LsgYjuS/edit?usp=sharing&ouid=115340390490868262616&rtpof=true&sd=true"
        st.info("âœ… ä½¿ç”¨é è¨­ç­è¡¨é€£çµ")
        submit_button = True
    
    if submit_button:
        if not main_sheet_url.strip():
            st.error("âŒ è«‹è¼¸å…¥å“¡å·¥ç­è¡¨çš„ Google Sheets é€£çµ")
            return
        
        if '/d/' not in main_sheet_url:
            st.error("âŒ Google Sheets é€£çµæ ¼å¼ä¸æ­£ç¢ºï¼Œè«‹ç¢ºä¿åŒ…å«å®Œæ•´çš„åˆ†äº«é€£çµ")
            return
        
        with st.spinner("ğŸ”„ æ­£åœ¨è¼‰å…¥ç­è¡¨è³‡æ–™..."):
            df, shift_dict, message = load_data_from_urls(main_sheet_url)
        
        if df is not None:
            # æ›´æ–° session state
            st.session_state.df = df
            st.session_state.shift_dict = shift_dict
            
            # å–å¾—äººäº‹è™Ÿæ¸…å–®
            personnel_numbers = df.iloc[1, :].tolist()
            personnel_count = sum(1 for num in personnel_numbers if pd.notna(num) and str(num).strip())
            
            st.success("âœ… ç­è¡¨è¼‰å…¥å®Œæˆï¼")
            st.info(f"ğŸ“Š ç­è¡¨å°ºå¯¸: {df.shape}")
            st.info(f"ğŸ‘¥ æ‰¾åˆ° {personnel_count} å€‹äººäº‹è™Ÿ")
            st.info(f"ğŸ”¢ ç­ç¨®æ•¸é‡: {len(shift_dict)} ç¨®")
            
            # æ¸…é™¤å¿«å–
            st.cache_data.clear()
            
        else:
            st.error(message)

def query_page():
    """æŸ¥è©¢é é¢"""
    st.header("ğŸ” å“¡å·¥åŠ ç­æ™‚æ•¸æŸ¥è©¢")
    
    if st.session_state.df is None:
        st.warning("âš ï¸ è«‹å…ˆè¼‰å…¥ç­è¡¨è³‡æ–™")
        return
    
    df = st.session_state.df
    
    # å–å¾—äººäº‹è™Ÿæ¸…å–®
    personnel_numbers = df.iloc[1, :].tolist()
    personnel_options = []
    
    for i, num in enumerate(personnel_numbers):
        if pd.notna(num) and str(num).strip():
            col_name = chr(65 + i) if i < 26 else chr(65 + i//26 - 1) + chr(65 + i%26)
            personnel_options.append(f"{num} (Column {col_name})")
    
    if not personnel_options:
        st.error("âŒ æœªæ‰¾åˆ°ä»»ä½•äººäº‹è™Ÿ")
        return
    
    with st.form("query_form"):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            selected_personnel = st.selectbox("é¸æ“‡äººäº‹è™Ÿ", personnel_options)
        
        with col2:
            year = st.number_input("è¥¿å…ƒå¹´", min_value=2020, max_value=2030, 
                                 value=datetime.now().year)
        
        with col3:
            month = st.selectbox("æœˆä»½", 
                               [(i, f"{i}æœˆ") for i in range(1, 13)], 
                               index=datetime.now().month-1,
                               format_func=lambda x: x[1])
        
        col_query, col_preview = st.columns(2)
        with col_query:
            submit_query = st.form_submit_button("ğŸ” æŸ¥è©¢åŠ ç­æ™‚æ•¸", type="primary")
        with col_preview:
            preview_schedule = st.form_submit_button("ğŸ‘ï¸ é è¦½ç­è¡¨", type="secondary")
    
    # è™•ç†ç­è¡¨é è¦½
    if preview_schedule:
        target_personnel = selected_personnel.split(' (')[0]
        month_value = month[0]
        
        # æŸ¥æ‰¾åŒ¹é…çš„æ¬„ä½
        personnel_numbers = df.iloc[1, :].tolist()
        matching_columns = []
        
        for col_idx, personnel_num in enumerate(personnel_numbers):
            if pd.notna(personnel_num) and str(personnel_num).strip() == target_personnel:
                matching_columns.append(col_idx)
        
        if matching_columns:
            with st.spinner(f"ğŸ‘ï¸ æ­£åœ¨ç”Ÿæˆ {target_personnel} çš„ {year}å¹´{month_value}æœˆ ç­è¡¨é è¦½..."):
                preview_data = generate_schedule_preview(target_personnel, year, month_value, matching_columns)
                st.session_state.preview_data = {
                    'personnel': target_personnel,
                    'year': year,
                    'month': month_value,
                    'data': preview_data
                }

    # é¡¯ç¤ºç­è¡¨é è¦½
    if st.session_state.preview_data is not None:
        preview_info = st.session_state.preview_data
        st.subheader(f"ğŸ‘ï¸ {preview_info['personnel']} - {preview_info['year']}å¹´{preview_info['month']}æœˆç­è¡¨é è¦½")
        
        # é¡¯ç¤ºçµ±è¨ˆè³‡è¨Š
        total_days = len(preview_info['data'])
        work_days = sum(1 for item in preview_info['data'] if item['ç­æ¬¡'] != 'ä¼‘å‡')
        holiday_work_days = sum(1 for item in preview_info['data'] if item['ç­æ¬¡'] != 'ä¼‘å‡' and item['é¡å‹'] == 'å‡æ—¥')
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("ç¸½å¤©æ•¸", f"{total_days} å¤©")
        with col2:
            st.metric("ä¸Šç­å¤©æ•¸", f"{work_days} å¤©")
        with col3:
            st.metric("å‡æ—¥ä¸Šç­", f"{holiday_work_days} å¤©")
        
        # é¡¯ç¤ºè©³ç´°ç­è¡¨
        df_preview = pd.DataFrame(preview_info['data'])
        st.dataframe(df_preview, use_container_width=True)
    
    if submit_query:
        # æå–äººäº‹è™Ÿ
        target_personnel = selected_personnel.split(' (')[0]
        month_value = month[0]
        
        with st.spinner(f"ğŸ” æ­£åœ¨æŸ¥è©¢ {target_personnel} çš„ {year}å¹´{month_value}æœˆ åŠ ç­æ™‚æ•¸..."):
            # æŸ¥æ‰¾åŒ¹é…çš„æ¬„ä½
            personnel_numbers = df.iloc[1, :].tolist()
            matching_columns = []
            
            for col_idx, personnel_num in enumerate(personnel_numbers):
                if pd.notna(personnel_num) and str(personnel_num).strip() == target_personnel:
                    matching_columns.append(col_idx)
            
            if not matching_columns:
                st.error(f"âŒ æœªæ‰¾åˆ°äººäº‹è™Ÿ: {target_personnel}")
                return
            
            # è¨ˆç®—åŠ ç­æ™‚æ•¸
            weekday_hours, weekend_hours, total_hours, records, daily_breakdown = calculate_overtime_summary(
                target_personnel, year, month_value, matching_columns
            )
            
            # å„²å­˜æŸ¥è©¢çµæœ
            st.session_state.last_query_result = {
                'target_personnel': target_personnel,
                'year': year,
                'month': month_value,
                'matching_columns': matching_columns,
                'daily_breakdown': daily_breakdown,
                'weekday_hours': weekday_hours,
                'weekend_hours': weekend_hours,
                'total_hours': total_hours
            }
        
        # é¡¯ç¤ºçµæœ
        st.success("âœ… æŸ¥è©¢å®Œæˆï¼")
        
        # é¡¯ç¤ºè‡ªå®šç¾©å‡æ—¥è³‡è¨Š
        if st.session_state.custom_holidays:
            current_month_holidays = {k: v for k, v in st.session_state.custom_holidays.items()
                                    if k.startswith(f"{year}-{month_value:02d}-")}
            if current_month_holidays:
                st.info(f"ğŸ–ï¸ æœ¬æœˆè‡ªå®šç¾©å‡æ—¥ ({len(current_month_holidays)} å¤©)")
                holiday_list = []
                for date_key, desc in sorted(current_month_holidays.items()):
                    holiday_list.append(f"â€¢ {date_key}: {desc}")
                st.markdown("\n".join(holiday_list))
        
        # çµ±è¨ˆçµæœå¡ç‰‡
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("å¹³æ—¥åŠ ç­æ™‚æ•¸", f"{weekday_hours:.1f} å°æ™‚")
        with col2:
            st.metric("å‡æ—¥åŠ ç­æ™‚æ•¸", f"{weekend_hours:.1f} å°æ™‚")
        with col3:
            st.metric("ç¸½åŠ ç­æ™‚æ•¸", f"{total_hours:.1f} å°æ™‚")
        
        # è©³ç´°æ¯æ—¥è³‡æ–™
        if daily_breakdown:
            st.subheader("ğŸ“… è©³ç´°æ¯æ—¥åŠ ç­è¨˜éŒ„")
            
            # å‰µå»ºè¡¨æ ¼æ•¸æ“š
            table_data = []
            for date_str, hours in sorted(daily_breakdown.items()):
                if hours > 0:
                    date_parts = date_str.split('/')
                    check_year = int(date_parts[0])
                    check_month = int(date_parts[1])
                    check_day = int(date_parts[2])
                    day_type, is_weekend = get_day_type(check_year, check_month, check_day)
                    
                    table_data.append({
                        'æ—¥æœŸ': date_str,
                        'æ˜ŸæœŸ': day_type,
                        'åŠ ç­æ™‚æ•¸': f"{hours:.1f}å°æ™‚",
                        'é¡å‹': 'å‡æ—¥' if is_weekend else 'å¹³æ—¥'
                    })
            
            if table_data:
                df_display = pd.DataFrame(table_data)
                st.dataframe(df_display, use_container_width=True)
    
    # Excel åŒ¯å‡ºæŒ‰éˆ•
    if st.session_state.last_query_result is not None:
        st.subheader("ğŸ“Š åŒ¯å‡ºå ±è¡¨")
        
        # å‰µå»ºä¸‹è¼‰æŒ‰éˆ•çš„å®¹å™¨
        export_container = st.container()
        
        with export_container:
            if st.button("ğŸ“Š ç”¢ç”Ÿä¸¦ä¸‹è¼‰Excelå ±è¡¨", type="secondary", key="export_excel_btn"):
                with st.spinner("ğŸ“Š æ­£åœ¨ç”¢ç”ŸExcelå ±è¡¨..."):
                    result = st.session_state.last_query_result
                    success, file_content_or_error, weekday_total, weekend_total, total_hours_export, row_count = export_to_excel(
                        result['target_personnel'],
                        result['year'],
                        result['month'],
                        result['matching_columns'],
                        result['daily_breakdown']
                    )
                    
                    if success:
                        filename = f"{result['target_personnel']}_{result['year']}å¹´{result['month']:02d}æœˆ_åŠ ç­æ™‚æ•¸çµ±è¨ˆ.xlsx"
                        
                        st.success("âœ… Excelå ±è¡¨ç”¢ç”ŸæˆåŠŸï¼")
                        st.info(f"ğŸ“‹ åŒ…å«è³‡æ–™: {row_count} å¤©")
                        
                        # æä¾›ä¸‹è¼‰æŒ‰éˆ•
                        st.download_button(
                            label="ğŸ“¥ ä¸‹è¼‰Excelæª”æ¡ˆ",
                            data=file_content_or_error.getvalue(),
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_excel_btn"
                        )
                    else:
                        st.error(f"âŒ ExcelåŒ¯å‡ºå¤±æ•—: {file_content_or_error}")

def holiday_management_page():
    """è‡ªå®šç¾©å‡æ—¥ç®¡ç†é é¢"""
    st.header("ğŸ—“ï¸ è‡ªå®šç¾©å‡æ—¥ç®¡ç†")
    
    # æé†’ä½¿ç”¨è€…æ²’æœ‰è¨˜æ†¶åŠŸèƒ½
    st.warning("âš ï¸ æ³¨æ„ï¼šè‡ªå®šç¾©å‡æ—¥è¨­å®šåœ¨é—œé–‰ç€è¦½å™¨æˆ–é‡æ–°è¼‰å…¥é é¢å¾Œå°‡æœƒæ¸…é™¤ï¼Œæ²’æœ‰è¨˜æ†¶åŠŸèƒ½ã€‚")
    
    # æ–°å¢å‡æ—¥å€åŸŸ
    st.subheader("â• æ–°å¢è‡ªå®šç¾©å‡æ—¥")
    
    with st.form("add_holiday_form"):
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            holiday_year = st.number_input("å¹´ä»½", min_value=2020, max_value=2030, 
                                         value=datetime.now().year)
        
        with col2:
            holiday_month = st.selectbox("æœˆä»½", 
                                       [(i, f"{i}æœˆ") for i in range(1, 13)], 
                                       index=datetime.now().month-1,
                                       format_func=lambda x: x[1])
        
        with col3:
            holiday_day = st.number_input("æ—¥æœŸ", min_value=1, max_value=31, value=1)
        
        with col4:
            holiday_reason = st.text_input("å‡æ—¥åŸå› ", value="è‡ªå®šç¾©å‡æ—¥")
        
        col_add, col_remove = st.columns(2)
        with col_add:
            add_holiday = st.form_submit_button("â• æ–°å¢å‡æ—¥", type="primary")
        with col_remove:
            remove_holiday = st.form_submit_button("âŒ ç§»é™¤å‡æ—¥", type="secondary")
    
    # è™•ç†æ–°å¢å‡æ—¥
    if add_holiday:
        year_val = holiday_year
        month_val = holiday_month[0]
        day_val = holiday_day
        reason = holiday_reason.strip() if holiday_reason.strip() else "è‡ªå®šç¾©å‡æ—¥"
        
        try:
            # é©—è­‰æ—¥æœŸ
            test_date = date(year_val, month_val, day_val)
            date_key = f"{year_val}-{month_val:02d}-{day_val:02d}"
            
            weekdays = ['ä¸€', 'äºŒ', 'ä¸‰', 'å››', 'äº”', 'å…­', 'æ—¥']
            weekday = weekdays[test_date.weekday()]
            
            st.session_state.custom_holidays[date_key] = f"{reason}({weekday})"
            st.success(f"âœ… å·²æ–°å¢è‡ªå®šç¾©å‡æ—¥: {date_key} {reason}({weekday})")
            
        except ValueError:
            st.error(f"âŒ ç„¡æ•ˆæ—¥æœŸ: {year_val}-{month_val:02d}-{day_val:02d}")
    
    # è™•ç†ç§»é™¤å‡æ—¥
    if remove_holiday:
        year_val = holiday_year
        month_val = holiday_month[0]
        day_val = holiday_day
        date_key = f"{year_val}-{month_val:02d}-{day_val:02d}"
        
        if date_key in st.session_state.custom_holidays:
            removed = st.session_state.custom_holidays.pop(date_key)
            st.success(f"âœ… å·²ç§»é™¤è‡ªå®šç¾©å‡æ—¥: {date_key} ({removed})")
        else:
            st.warning(f"âš ï¸ è©²æ—¥æœŸä¸æ˜¯è‡ªå®šç¾©å‡æ—¥: {date_key}")
    
    # ç®¡ç†ç¾æœ‰å‡æ—¥
    st.subheader("ğŸ“… ç›®å‰è¨­å®šçš„è‡ªå®šç¾©å‡æ—¥")
    
    if st.button("ğŸ—‘ï¸ æ¸…é™¤æ‰€æœ‰å‡æ—¥", type="secondary"):
        if st.session_state.custom_holidays:
            st.session_state.custom_holidays.clear()
            st.success("âœ… å·²æ¸…é™¤æ‰€æœ‰è‡ªå®šç¾©å‡æ—¥")
        else:
            st.info("ğŸ“… ç›®å‰æ²’æœ‰è¨­å®šä»»ä½•è‡ªå®šç¾©å‡æ—¥")
    
    # é¡¯ç¤ºå‡æ—¥æ¸…å–®
    if st.session_state.custom_holidays:
        st.write(f"**ç›®å‰å…±æœ‰ {len(st.session_state.custom_holidays)} å€‹è‡ªå®šç¾©å‡æ—¥:**")
        
        # è½‰æ›ç‚ºè¡¨æ ¼é¡¯ç¤º
        holiday_data = []
        for date_key, description in sorted(st.session_state.custom_holidays.items()):
            holiday_data.append({
                'æ—¥æœŸ': date_key,
                'æè¿°': description,
                'å¹´': date_key.split('-')[0],
                'æœˆ': date_key.split('-')[1],
                'æ—¥': date_key.split('-')[2]
            })
        
        df_holidays = pd.DataFrame(holiday_data)
        st.dataframe(df_holidays[['æ—¥æœŸ', 'æè¿°']], use_container_width=True)
        
        # æŒ‰æœˆä»½åˆ†çµ„é¡¯ç¤º
        st.subheader("ğŸ“Š æŒ‰æœˆä»½åˆ†çµ„")
        months_dict = {}
        for date_key, description in st.session_state.custom_holidays.items():
            year_month = date_key[:7]  # YYYY-MM
            if year_month not in months_dict:
                months_dict[year_month] = []
            months_dict[year_month].append(f"{date_key}: {description}")
        
        for year_month, holidays in sorted(months_dict.items()):
            with st.expander(f"ğŸ“… {year_month} ({len(holidays)} å¤©)"):
                for holiday in holidays:
                    st.write(f"â€¢ {holiday}")
    else:
        st.info("ğŸ“… ç›®å‰æ²’æœ‰è¨­å®šä»»ä½•è‡ªå®šç¾©å‡æ—¥")

if __name__ == "__main__":
    main()