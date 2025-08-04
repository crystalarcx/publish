# 互動式員工班表加班時數統計系統 (Streamlit版) - 全面優化版 + 手動編輯班次功能
"""
員工班表加班時數統計系統
====================

功能特色：
- 從 Google Sheets 載入班表資料
- 自動計算員工加班時數
- 支援自定義假日設定
- 生成詳細的 Excel 報表
- 班表預覽功能
- 手動編輯班次功能

作者: AI Assistant
版本: 2.2 (新增手動編輯班次功能) - 修復版
"""

import pandas as pd
from datetime import datetime, date, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import streamlit as st
import warnings
import calendar
import io
import base64
import re
from typing import Dict, List, Tuple, Optional, Any, Union
from dataclasses import dataclass
import time

warnings.filterwarnings('ignore')

# ===== 設定常數 =====
class Config:
    """系統設定常數"""
    # 班表相關設定
    DEFAULT_SHIFT_SHEET_URL = "https://docs.google.com/spreadsheets/d/1JfhGZYRBWj6yp90o-sA0DrhzkcEM1Wfd_vqiEZEYd5c/edit?usp=sharing"
    DEFAULT_MAIN_SHEET_URL = "https://docs.google.com/spreadsheets/d/1U8qLraVCRKJhySk0y93I_POP_LsgYjuS/edit?usp=sharing&ouid=115340390490868262616&rtpof=true&sd=true"
    OVERTIME_FORM_URL = "https://docs.google.com/document/d/1T75rw_3hQtIaBTGMFxa09G93Atihf4h-883Kg1tqPpo/edit?usp=sharing"
    
    # 指定的人事號清單
    ALLOWED_PERSONNEL = ['A30825', 'A408J6', 'A40837', 'A608Q2', 'A50847', 'A60811', 'A708J6', 'A808L5', 'B00505', 'A81205', 'A908H8']
    
    # 班表範圍設定
    MAX_ROWS = 36
    MAX_COLS = 83
    
    # 加班時數相關設定
    MAX_WEEKDAY_HOURS = 46.0
    AUTO_ADD_HOURS = 2.0
    WEEKEND_MIN_HOURS_THRESHOLD = 3.0
    EARLY_MORNING_CUTOFF = 5  # 05:00
    
    # 日期相關設定
    MIN_YEAR = 2020
    MAX_YEAR = 2030
    
    # 優先日期設定（週二、週四優先）
    HIGH_PRIORITY_WEEKDAYS = [1, 3]  # 週二、週四
    MEDIUM_PRIORITY_WEEKDAYS = [0, 2, 4]  # 週一、週三、週五

@dataclass
class ShiftInfo:
    """班次資訊資料類別"""
    shift_type: str
    overtime_hours_1: Optional[str]
    overtime_hours_2: Optional[str]
    cross_day_hours: Optional[str]

@dataclass
class QueryResult:
    """查詢結果資料類別"""
    target_personnel: str
    year: int
    month: int
    matching_columns: List[int]
    daily_breakdown: Dict[str, float]
    weekday_hours: float
    weekend_hours: float
    total_hours: float

@dataclass
class PreviewData:
    """預覽資料類別"""
    personnel: str
    year: int
    month: int
    data: List[Dict[str, Any]]
    editable: bool = False  # 新增：是否可編輯標記

# ===== Streamlit 頁面配置 =====
st.set_page_config(
    page_title="員工班表加班時數統計系統",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ===== Session State 管理 =====
class SessionStateManager:
    """Session State 管理類別"""
    
    @staticmethod
    def initialize():
        """初始化所有 session state"""
        default_states = {
            'df': None,
            'shift_dict': {},
            'custom_holidays': {},
            'last_query_result': None,
            'current_page': "載入班表資料",
            'preview_data': None,
            'data_load_time': None,
            'cache_version': 0,
            'manual_shifts': {},  # 新增：手動修改的班次資料 {personnel_year_month: {date: shift}}
            'editing_mode': False,  # 新增：編輯模式標記
            'current_edit_key': None,  # 新增：當前編輯的key
        }
        
        for key, default_value in default_states.items():
            if key not in st.session_state:
                st.session_state[key] = default_value
    
    @staticmethod
    def clear_cache():
        """清除快取並更新版本號"""
        st.cache_data.clear()
        st.session_state.cache_version += 1
        st.session_state.data_load_time = datetime.now()
    
    @staticmethod
    def get_manual_shift_key(personnel: str, year: int, month: int) -> str:
        """生成手動班次的key"""
        return f"{personnel}_{year}_{month:02d}"
    
    @staticmethod
    def get_manual_shift(personnel: str, year: int, month: int, day: int) -> Optional[str]:
        """取得手動設定的班次"""
        key = SessionStateManager.get_manual_shift_key(personnel, year, month)
        if key in st.session_state.manual_shifts:
            date_str = f"{year}/{month:02d}/{day:02d}"
            if date_str in st.session_state.manual_shifts[key]:
                return st.session_state.manual_shifts[key][date_str]
        return None
    
    @staticmethod
    def set_manual_shift(personnel: str, year: int, month: int, day: int, shift: str):
        """設定手動班次"""
        key = SessionStateManager.get_manual_shift_key(personnel, year, month)
        if key not in st.session_state.manual_shifts:
            st.session_state.manual_shifts[key] = {}
        
        date_str = f"{year}/{month:02d}/{day:02d}"
        if shift.strip():
            # 設定新的班次
            st.session_state.manual_shifts[key][date_str] = shift.strip()
        else:
            # 如果設為空，記錄為空班次（表示手動設為休假）
            st.session_state.manual_shifts[key][date_str] = ""

# ===== 工具函數類別 =====
class DataLoader:
    """資料載入相關功能"""
    
    @staticmethod
    def convert_google_sheet_url(url: str) -> Optional[str]:
        """
        將 Google Sheets URL 轉換為可直接讀取的 CSV URL
        
        Args:
            url: Google Sheets 分享連結
            
        Returns:
            CSV 格式的下載連結，如果格式不正確則返回 None
        """
        if not url or '/d/' not in url:
            return None
            
        try:
            sheet_id = url.split('/d/')[1].split('/')[0]
            return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
        except (IndexError, AttributeError):
            return None
    
    @staticmethod
    def validate_url_format(url: str) -> Tuple[bool, str]:
        """
        驗證 Google Sheets URL 格式
        
        Args:
            url: 要驗證的 URL
            
        Returns:
            (是否有效, 錯誤訊息)
        """
        if not url or not url.strip():
            return False, "URL 不能為空"
        
        if '/d/' not in url:
            return False, "URL 格式不正確，請確保包含 Google Sheets 的完整分享連結"
        
        if 'docs.google.com/spreadsheets' not in url:
            return False, "請提供有效的 Google Sheets 連結"
        
        return True, ""
    
    @staticmethod
    @st.cache_data(ttl=300)  # 快取 5 分鐘
    def load_data_from_urls(main_sheet_url: str, cache_version: int = 0) -> Tuple[Optional[pd.DataFrame], Optional[Dict], str]:
        """
        從 URL 載入資料（帶快取功能）
        
        Args:
            main_sheet_url: 主要班表 URL
            cache_version: 快取版本號（用於強制更新快取）
            
        Returns:
            (DataFrame, 班次字典, 狀態訊息)
        """
        try:
            # 進度條
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # 驗證 URL
            is_valid, error_msg = DataLoader.validate_url_format(main_sheet_url)
            if not is_valid:
                return None, None, f"❌ URL 驗證失敗: {error_msg}"
            
            status_text.text("🔄 正在轉換 URL...")
            progress_bar.progress(10)
            
            main_csv_url = DataLoader.convert_google_sheet_url(main_sheet_url)
            shift_csv_url = DataLoader.convert_google_sheet_url(Config.DEFAULT_SHIFT_SHEET_URL)
            
            if not main_csv_url or not shift_csv_url:
                return None, None, "❌ URL 轉換失敗"
            
            status_text.text("📊 正在讀取員工班表...")
            progress_bar.progress(30)
            
            # 讀取員工班表
            df_full = pd.read_csv(main_csv_url)
            df = df_full.iloc[:Config.MAX_ROWS, :Config.MAX_COLS]  # 選取指定範圍
            
            status_text.text("🔢 正在讀取班種對照表...")
            progress_bar.progress(60)
            
            # 讀取班種對照表
            shift_df = pd.read_csv(shift_csv_url)
            
            status_text.text("🔨 正在建立班種字典...")
            progress_bar.progress(80)
            
            # 建立班種字典
            shift_dict = DataProcessor.build_shift_dictionary(shift_df)
            
            status_text.text("✅ 資料載入完成！")
            progress_bar.progress(100)
            
            # 清理進度顯示
            time.sleep(0.5)
            progress_bar.empty()
            status_text.empty()
            
            # 資料驗證
            personnel_count = DataValidator.count_allowed_personnel(df)
            
            return df, shift_dict, f"✅ 資料讀取成功！班表: {df.shape}, 指定人員: {personnel_count} 人"
            
        except pd.errors.EmptyDataError:
            return None, None, "❌ 資料檔案為空或格式不正確"
        except pd.errors.ParserError as e:
            return None, None, f"❌ 資料解析失敗: 檔案格式可能有問題"
        except Exception as e:
            return None, None, f"❌ 資料讀取失敗: {str(e)}"

class DataProcessor:
    """資料處理相關功能"""
    
    @staticmethod
    def build_shift_dictionary(shift_df: pd.DataFrame) -> Dict[str, ShiftInfo]:
        """
        建立班種字典
        
        Args:
            shift_df: 班種對照表 DataFrame
            
        Returns:
            班種字典
        """
        shift_dict = {}
        
        for index, row in shift_df.iterrows():
            try:
                shift_type = str(row.iloc[0]).strip()
                if not shift_type or shift_type == 'nan':
                    continue
                    
                overtime_hours_1 = row.iloc[1] if len(row) > 1 else None
                overtime_hours_2 = row.iloc[2] if len(row) > 2 else None
                cross_day_hours = row.iloc[3] if len(row) > 3 else None
                
                shift_dict[shift_type] = ShiftInfo(
                    shift_type=shift_type,
                    overtime_hours_1=overtime_hours_1,
                    overtime_hours_2=overtime_hours_2,
                    cross_day_hours=cross_day_hours
                )
            except (IndexError, ValueError) as e:
                st.warning(f"⚠️ 班種資料第 {index+1} 行格式異常，已跳過")
                continue
        
        return shift_dict
    
    @staticmethod
    def find_matching_personnel_columns(df: pd.DataFrame, target_personnel: str) -> List[int]:
        """
        查找匹配的人事號欄位
        
        Args:
            df: 班表 DataFrame
            target_personnel: 目標人事號
            
        Returns:
            匹配的欄位索引列表
        """
        personnel_numbers = df.iloc[1, :].tolist()
        matching_columns = []
        
        for col_idx, personnel_num in enumerate(personnel_numbers):
            if pd.notna(personnel_num) and str(personnel_num).strip() == target_personnel:
                matching_columns.append(col_idx)
        
        return matching_columns
    
    @staticmethod
    def get_personnel_options(df: pd.DataFrame) -> List[str]:
        """
        取得指定人事號選項列表
        
        Args:
            df: 班表 DataFrame
            
        Returns:
            指定人事號選項列表
        """
        personnel_numbers = df.iloc[1, :].tolist()
        personnel_options = []
        
        for i, num in enumerate(personnel_numbers):
            if pd.notna(num) and str(num).strip() in Config.ALLOWED_PERSONNEL:
                col_name = DataProcessor.get_column_name(i)
                personnel_options.append(f"{num} (Column {col_name})")
        
        return personnel_options
    
    @staticmethod
    def get_column_name(index: int) -> str:
        """
        將欄位索引轉換為 Excel 欄位名稱 (A, B, C, ...)
        
        Args:
            index: 欄位索引
            
        Returns:
            Excel 欄位名稱
        """
        if index < 26:
            return chr(65 + index)
        else:
            return chr(65 + index//26 - 1) + chr(65 + index%26)
    
    @staticmethod
    def get_effective_shift(df: pd.DataFrame, personnel: str, year: int, month: int, day: int, matching_columns: List[int]) -> str:
        """
        取得有效的班次（優先使用手動設定，否則使用原始班次）
        修復版：正確處理空值和 NaN
        
        Args:
            df: 班表 DataFrame
            personnel: 人事號
            year: 年份
            month: 月份
            day: 日期
            matching_columns: 匹配的欄位列表
            
        Returns:
            有效的班次（空字串表示休假）
        """
        # 優先檢查手動設定的班次
        manual_shift = SessionStateManager.get_manual_shift(personnel, year, month, day)
        if manual_shift is not None:
            return manual_shift  # 可能是空字串（表示手動設為休假）
        
        # 使用原始班次
        for col_idx in matching_columns:
            column_data = df.iloc[:, col_idx]
            row_idx = day + 2
            
            if row_idx < len(column_data):
                value = column_data.iloc[row_idx]
                
                # 更嚴格的空值檢查
                if pd.isna(value) or value is None:
                    continue
                    
                shift_value = str(value).strip()
                
                # 檢查是否為有效的班次值
                if shift_value and shift_value.lower() not in ['nan', 'none', '']:
                    return shift_value
        
        return ""  # 沒有找到有效班次，返回空字串表示休假

class TimeCalculator:
    """時間計算相關功能"""
    
    @staticmethod
    def calculate_hours(time_range: Union[str, float, None]) -> Optional[float]:
        """
        計算時間範圍的小時數（優化版）
        
        Args:
            time_range: 時間範圍字串或數值
            
        Returns:
            計算出的小時數，無法計算則返回 None
        """
        if not time_range or pd.isna(time_range):
            return None

        time_str = str(time_range).strip()

        # 處理純數字（小時數）
        if TimeCalculator._is_pure_number(time_str):
            try:
                hours = float(time_str.replace(',', '.'))
                return hours if 0 <= hours <= 24 else None
            except ValueError:
                pass

        # 處理時間範圍格式
        if '-' not in time_str:
            return None

        return TimeCalculator._parse_time_range(time_str)
    
    @staticmethod
    def _is_pure_number(time_str: str) -> bool:
        """檢查是否為純數字"""
        # 移除常見的分隔符號
        cleaned = time_str.replace(',', '.').replace(' ', '')
        try:
            float(cleaned)
            return '-' not in time_str
        except ValueError:
            return False
    
    @staticmethod
    def _parse_time_range(time_str: str) -> Optional[float]:
        """解析時間範圍字串"""
        try:
            # 清理時間字串
            time_str = time_str.replace(' ', '').replace(',', '')
            
            parts = time_str.split('-')
            if len(parts) != 2:
                return None
            
            start_str, end_str = parts
            
            # 解析開始和結束時間
            start_hour, start_min = TimeCalculator._parse_time_component(start_str)
            end_hour, end_min = TimeCalculator._parse_time_component(end_str)
            
            if start_hour is None or end_hour is None:
                return None
            
            # 轉換為分鐘並計算時差
            start_minutes = start_hour * 60 + start_min
            end_minutes = end_hour * 60 + end_min
            
            # 處理跨日情況
            if end_minutes <= start_minutes:
                end_minutes += 24 * 60
            
            # 計算小時數
            total_minutes = end_minutes - start_minutes
            hours = total_minutes / 60
            
            return hours if hours > 0 else None
            
        except Exception:
            return None
    
    @staticmethod
    def _parse_time_component(time_str: str) -> Tuple[Optional[int], Optional[int]]:
        """
        解析單個時間組件
        
        Args:
            time_str: 時間字串（如 "14:30", "1430", "14"）
            
        Returns:
            (小時, 分鐘) 或 (None, None)
        """
        time_str = time_str.strip()
        
        # HH:MM 格式
        if ':' in time_str:
            try:
                parts = time_str.split(':')
                if len(parts) == 2:
                    hour = int(parts[0])
                    minute = int(parts[1])
                    if 0 <= hour <= 23 and 0 <= minute <= 59:
                        return hour, minute
            except ValueError:
                pass
        
        # HHMM 格式
        if len(time_str) == 4 and time_str.isdigit():
            try:
                hour = int(time_str[:2])
                minute = int(time_str[2:])
                if 0 <= hour <= 23 and 0 <= minute <= 59:
                    return hour, minute
            except ValueError:
                pass
        
        # HH 格式
        if time_str.isdigit() and 1 <= len(time_str) <= 2:
            try:
                hour = int(time_str)
                if 0 <= hour <= 23:
                    return hour, 0
            except ValueError:
                pass
        
        # 小數點格式
        try:
            hour_decimal = float(time_str)
            if 0 <= hour_decimal <= 24:
                hour = int(hour_decimal)
                minute = int((hour_decimal - hour) * 60)
                if 0 <= hour <= 23 and 0 <= minute <= 59:
                    return hour, minute
        except ValueError:
            pass
        
        return None, None

class DateHelper:
    """日期處理相關功能"""
    
    @staticmethod
    def get_day_type(year: int, month: int, day: int) -> Tuple[str, bool]:
        """
        判斷日期是平日還是假日（含自定義假日）
        
        Args:
            year: 年份
            month: 月份
            day: 日期
            
        Returns:
            (日期類型描述, 是否為假日)
        """
        try:
            # 檢查自定義假日
            date_key = f"{year}-{month:02d}-{day:02d}"
            if date_key in st.session_state.custom_holidays:
                return st.session_state.custom_holidays[date_key], True
            
            # 一般週末判斷
            current_date = date(year, month, day)
            weekday = current_date.weekday()
            
            if weekday == 5:  # 星期六
                return "假日(六)", True
            elif weekday == 6:  # 星期日
                return "假日(日)", True
            else:  # 平日
                weekdays = ["一", "二", "三", "四", "五"]
                return f"平日({weekdays[weekday]})", False
                
        except ValueError:
            return "無效日期", False
    
    @staticmethod
    def get_month_date_range(year: int, month: int) -> List[int]:
        """
        取得指定月份的所有日期
        
        Args:
            year: 年份
            month: 月份
            
        Returns:
            該月份的所有日期列表
        """
        try:
            _, last_day = calendar.monthrange(year, month)
            return list(range(1, last_day + 1))
        except ValueError:
            return list(range(1, 32))  # 備用方案

class DataValidator:
    """資料驗證相關功能"""
    
    @staticmethod
    def count_allowed_personnel(df: pd.DataFrame) -> int:
        """計算指定的人事號數量"""
        if df is None or df.empty:
            return 0
        
        personnel_numbers = df.iloc[1, :].tolist()
        return sum(1 for num in personnel_numbers 
                  if pd.notna(num) and str(num).strip() in Config.ALLOWED_PERSONNEL)
    
    @staticmethod
    def validate_query_parameters(personnel: str, year: int, month: int) -> Tuple[bool, str]:
        """
        驗證查詢參數
        
        Args:
            personnel: 人事號
            year: 年份
            month: 月份
            
        Returns:
            (是否有效, 錯誤訊息)
        """
        if not personnel or not personnel.strip():
            return False, "請選擇人事號"
        
        if not (Config.MIN_YEAR <= year <= Config.MAX_YEAR):
            return False, f"年份必須在 {Config.MIN_YEAR} 到 {Config.MAX_YEAR} 之間"
        
        if not (1 <= month <= 12):
            return False, "月份必須在 1 到 12 之間"
        
        return True, ""

class OvertimeCalculator:
    """加班時數計算功能"""
    
    @staticmethod
    def calculate_overtime_summary(target_personnel: str, year: int, month: int, matching_columns: List[int]) -> QueryResult:
        """
        計算指定人員的加班時數統計（支援手動班次）
        
        Args:
            target_personnel: 目標人事號
            year: 年份
            month: 月份
            matching_columns: 匹配的欄位列表
            
        Returns:
            查詢結果物件
        """
        df = st.session_state.df
        shift_dict = st.session_state.shift_dict
        
        # 初始化變數
        daily_records = []
        cross_day_records = defaultdict(float)
        worked_weekdays = set()
        
        # 收集所有班次資料（優先使用手動設定的班次）
        for day in DateHelper.get_month_date_range(year, month):
            try:
                current_date = date(year, month, day)
                date_str = f"{year}/{month:02d}/{day:02d}"
                day_type, is_weekend = DateHelper.get_day_type(year, month, day)
                
                # 取得有效班次（手動或原始）
                effective_shift = DataProcessor.get_effective_shift(
                    df, target_personnel, year, month, day, matching_columns
                )
                
                # 記錄有上班的平日
                if effective_shift and not is_weekend:
                    worked_weekdays.add(date_str)
                
                # 處理班次資料
                if effective_shift in shift_dict and effective_shift:
                    overtime_data = OvertimeCalculator._calculate_daily_overtime(
                        shift_dict[effective_shift], current_date, date_str, day_type, is_weekend
                    )
                    
                    if overtime_data:
                        daily_records.append(overtime_data)
                        
                        # 處理跨天時數
                        if overtime_data['cross_day_overtime'] > 0:
                            next_date = current_date + timedelta(days=1)
                            next_date_str = f"{next_date.year}/{next_date.month:02d}/{next_date.day:02d}"
                            cross_day_records[next_date_str] += overtime_data['cross_day_overtime']
            
            except ValueError:
                continue
        
        # 建立每日加班時數統計
        final_daily_overtime = OvertimeCalculator._build_daily_overtime_summary(daily_records, cross_day_records)
        
        # 計算平日和假日時數
        weekday_hours, weekend_hours = OvertimeCalculator._calculate_weekday_weekend_hours(final_daily_overtime, year, month)
        
        # 調整平日時數（46小時限制和自動補足）
        final_daily_overtime, weekday_hours = OvertimeCalculator._adjust_weekday_hours(
            final_daily_overtime, weekday_hours, worked_weekdays, year, month
        )
        
        total_hours = weekday_hours + weekend_hours
        
        return QueryResult(
            target_personnel=target_personnel,
            year=year,
            month=month,
            matching_columns=matching_columns,
            daily_breakdown=dict(final_daily_overtime),
            weekday_hours=weekday_hours,
            weekend_hours=weekend_hours,
            total_hours=total_hours
        )
    
    @staticmethod
    def _calculate_daily_overtime(shift_info: ShiftInfo, current_date: date, date_str: str, day_type: str, is_weekend: bool) -> Optional[Dict]:
        """計算單日加班時數"""
        current_day_overtime = 0.0
        next_day_overtime = 0.0
        
        # 計算當天加班時數
        if pd.notna(shift_info.overtime_hours_1) and str(shift_info.overtime_hours_1).strip():
            hours_1 = TimeCalculator.calculate_hours(str(shift_info.overtime_hours_1))
            if hours_1:
                current_day_overtime += hours_1
        
        if pd.notna(shift_info.overtime_hours_2) and str(shift_info.overtime_hours_2).strip():
            hours_2 = TimeCalculator.calculate_hours(str(shift_info.overtime_hours_2))
            if hours_2:
                current_day_overtime += hours_2
        
        # 計算跨天時數
        if pd.notna(shift_info.cross_day_hours) and str(shift_info.cross_day_hours).strip():
            cross_hours = TimeCalculator.calculate_hours(str(shift_info.cross_day_hours))
            if cross_hours:
                next_day_overtime = cross_hours
        
        # 只有當有加班時數時才返回記錄
        if current_day_overtime > 0 or next_day_overtime > 0:
            return {
                'date': date_str,
                'day_type': day_type,
                'is_weekend': is_weekend,
                'shift': shift_info.shift_type,
                'current_day_overtime': current_day_overtime,
                'cross_day_overtime': next_day_overtime
            }
        
        return None
    
    @staticmethod
    def _build_daily_overtime_summary(daily_records: List[Dict], cross_day_records: Dict[str, float]) -> defaultdict:
        """建立每日加班時數統計"""
        final_daily_overtime = defaultdict(float)
        
        # 加入當天加班時數
        for record in daily_records:
            date_str = record['date']
            current_overtime = record['current_day_overtime']
            if current_overtime > 0:
                final_daily_overtime[date_str] += current_overtime
        
        # 加入跨天時數
        for date_str, cross_hours in cross_day_records.items():
            final_daily_overtime[date_str] += cross_hours
        
        return final_daily_overtime
    
    @staticmethod
    def _calculate_weekday_weekend_hours(final_daily_overtime: Dict[str, float], year: int, month: int) -> Tuple[float, float]:
        """計算平日和假日總時數"""
        weekday_hours = 0.0
        weekend_hours = 0.0
        
        for date_str, total_hours in final_daily_overtime.items():
            try:
                date_parts = date_str.split('/')
                check_year = int(date_parts[0])
                check_month = int(date_parts[1])
                check_day = int(date_parts[2])
                
                _, is_weekend = DateHelper.get_day_type(check_year, check_month, check_day)
                
                if is_weekend:
                    weekend_hours += total_hours
                else:
                    weekday_hours += total_hours
            
            except (ValueError, IndexError):
                continue
        
        return weekday_hours, weekend_hours
    
    @staticmethod
    def _adjust_weekday_hours(final_daily_overtime: defaultdict, weekday_hours: float, worked_weekdays: set, year: int, month: int) -> Tuple[defaultdict, float]:
        """調整平日加班時數（46小時限制和自動補足）"""
        # 超過46小時則減少
        if weekday_hours > Config.MAX_WEEKDAY_HOURS:
            final_daily_overtime, weekday_hours = OvertimeCalculator._reduce_excess_hours(
                final_daily_overtime, weekday_hours, year, month
            )
        
        # 少於46小時則自動補足
        elif weekday_hours < Config.MAX_WEEKDAY_HOURS:
            final_daily_overtime, weekday_hours = OvertimeCalculator._add_missing_hours(
                final_daily_overtime, weekday_hours, worked_weekdays, year, month
            )
        
        return final_daily_overtime, weekday_hours
    
    @staticmethod
    def _reduce_excess_hours(final_daily_overtime: defaultdict, weekday_hours: float, year: int, month: int) -> Tuple[defaultdict, float]:
        """減少超過46小時的部分"""
        # 收集平日的時數資料
        weekday_dates = []
        for date_str, hours in final_daily_overtime.items():
            if hours > 0:
                try:
                    date_parts = date_str.split('/')
                    check_year = int(date_parts[0])
                    check_month = int(date_parts[1])
                    check_day = int(date_parts[2])
                    
                    _, is_weekend = DateHelper.get_day_type(check_year, check_month, check_day)
                    
                    if not is_weekend:
                        weekday_dates.append((date_str, hours))
                except (ValueError, IndexError):
                    continue
        
        # 按時數排序，優先刪除較小的時數
        weekday_dates.sort(key=lambda x: x[1])
        
        excess_hours = weekday_hours - Config.MAX_WEEKDAY_HOURS
        removed_hours = 0.0
        
        for date_str, hours in weekday_dates:
            if removed_hours + hours <= excess_hours:
                # 完全移除這一天
                final_daily_overtime[date_str] = 0.0
                removed_hours += hours
                weekday_hours -= hours
                
                if removed_hours >= excess_hours:
                    break
            elif removed_hours < excess_hours:
                # 部分移除
                remaining_to_remove = excess_hours - removed_hours
                final_daily_overtime[date_str] -= remaining_to_remove
                weekday_hours -= remaining_to_remove
                break
        
        return final_daily_overtime, weekday_hours
    
    @staticmethod
    def _add_missing_hours(final_daily_overtime: defaultdict, weekday_hours: float, worked_weekdays: set, year: int, month: int) -> Tuple[defaultdict, float]:
        """自動補足平日加班時數到46小時"""
        shortage = Config.MAX_WEEKDAY_HOURS - weekday_hours
        
        # 找出可用的平日
        available_weekdays = []
        for day in DateHelper.get_month_date_range(year, month):
            try:
                check_date = date(year, month, day)
                date_str = f"{year}/{month:02d}/{day:02d}"
                day_type, is_weekend = DateHelper.get_day_type(year, month, day)
                weekday_num = check_date.weekday()
                
                if not is_weekend and date_str not in worked_weekdays:
                    # 設定優先級
                    if weekday_num in Config.HIGH_PRIORITY_WEEKDAYS:
                        priority = 1
                    elif weekday_num in Config.MEDIUM_PRIORITY_WEEKDAYS:
                        priority = 2
                    else:
                        priority = 3
                    
                    available_weekdays.append((date_str, day_type, weekday_num, priority))
            except ValueError:
                continue
        
        # 按優先順序排序
        available_weekdays.sort(key=lambda x: (x[3], x[0]))
        
        if available_weekdays:
            days_needed = int(shortage / Config.AUTO_ADD_HOURS) + (1 if shortage % Config.AUTO_ADD_HOURS > 0 else 0)
            
            for i, (date_str, day_type, weekday_num, priority) in enumerate(available_weekdays):
                if i < days_needed:
                    final_daily_overtime[date_str] += Config.AUTO_ADD_HOURS
                    weekday_hours += Config.AUTO_ADD_HOURS
        
        return final_daily_overtime, weekday_hours

class TextProcessor:
    """文字處理相關功能"""
    
    @staticmethod
    def extract_chinese_note(time_string: str) -> str:
        """
        從時間字串中提取中文註記
        
        Args:
            time_string: 時間字串
            
        Returns:
            提取的中文註記，預設為"臨床業務"
        """
        if not time_string:
            return "臨床業務"
        
        chinese_pattern = r'[\u4e00-\u9fff]+|\([^\)]*[\u4e00-\u9fff][^\)]*\)'
        chinese_matches = re.findall(chinese_pattern, time_string)
        
        if chinese_matches:
            chinese_note = chinese_matches[0]
            chinese_note = chinese_note.replace('(', '').replace(')', '')
            return chinese_note
        else:
            return "臨床業務"

class SchedulePreview:
    """班表預覽功能"""
    
    @staticmethod
    def generate_schedule_preview(target_personnel: str, year: int, month: int, matching_columns: List[int], editable: bool = False) -> PreviewData:
        """
        生成班表預覽資料（修復版）
        
        Args:
            target_personnel: 目標人事號
            year: 年份
            month: 月份
            matching_columns: 匹配的欄位列表
            editable: 是否為編輯模式
            
        Returns:
            預覽資料物件
        """
        df = st.session_state.df
        preview_data = []
        
        for day in DateHelper.get_month_date_range(year, month):
            try:
                current_date = date(year, month, day)
                date_str = f"{year}/{month:02d}/{day:02d}"
                day_type, is_weekend = DateHelper.get_day_type(year, month, day)
                
                # 取得有效班次（優先使用手動設定）
                effective_shift = DataProcessor.get_effective_shift(
                    df, target_personnel, year, month, day, matching_columns
                )
                
                # 檢查是否為手動修改的班次
                manual_shift = SessionStateManager.get_manual_shift(target_personnel, year, month, day)
                is_manual = manual_shift is not None
                
                # 正確處理班次顯示
                shift_display = effective_shift if effective_shift else '休假'
                
                preview_data.append({
                    '日期': f"{day:02d}",
                    '星期': day_type,
                    '班次': shift_display,
                    '類型': '假日' if is_weekend else '平日',
                    '手動修改': '✓' if is_manual else '',
                    'day': day,  # 用於編輯
                    'original_shift': "",  # 將在編輯時動態取得
                })
                
            except ValueError:
                continue
        
        return PreviewData(
            personnel=target_personnel,
            year=year,
            month=month,
            data=preview_data,
            editable=editable
        )

class ShiftEditor:
    """班次編輯功能（新增類別）"""
    
    @staticmethod
    def render_shift_editor(preview_data: PreviewData):
        """
        渲染班次編輯界面
        
        Args:
            preview_data: 預覽資料物件
        """
        st.subheader("✏️ 班次編輯模式")
        
        # 編輯說明
        st.info("💡 說明：在下方表格中直接修改班次，空白表示休假。修改會即時儲存到記憶體中。")
        
        # 取得可用班次選項
        available_shifts = ShiftEditor._get_available_shifts()
        shift_options = [''] + available_shifts  # 空白選項表示休假
        
        # 編輯功能按鈕
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("💾 儲存所有修改", type="primary"):
                st.success("✅ 修改已儲存！")
                st.session_state.editing_mode = False
                st.rerun()
        
        with col2:
            if st.button("↩️ 取消編輯", type="secondary"):
                st.session_state.editing_mode = False
                st.rerun()
        
        with col3:
            if st.button("🗑️ 清除本月修改", type="secondary"):
                ShiftEditor._clear_month_modifications(preview_data.personnel, preview_data.year, preview_data.month)
                st.success("✅ 已清除本月所有手動修改")
                st.rerun()
        
        st.markdown("---")
        
        # 渲染編輯表格
        ShiftEditor._render_edit_table(preview_data, shift_options)
        
        # 顯示修改統計
        ShiftEditor._render_modification_stats(preview_data)
    
    @staticmethod
    def _get_available_shifts() -> List[str]:
        """取得可用的班次選項"""
        if 'shift_dict' in st.session_state and st.session_state.shift_dict:
            return sorted(list(st.session_state.shift_dict.keys()))
        return []
    
    @staticmethod
    def _render_edit_table(preview_data: PreviewData, shift_options: List[str]):
        """渲染編輯表格（修復版）"""
        st.subheader("📝 班次編輯表格")
        
        df = st.session_state.df
        matching_columns = DataProcessor.find_matching_personnel_columns(df, preview_data.personnel)
        
        # 分週顯示
        weeks = ShiftEditor._group_days_by_week(preview_data.data, preview_data.year, preview_data.month)
        
        for week_num, week_data in weeks.items():
            with st.container():
                st.write(f"**第 {week_num} 週**")
                
                # 為每一週建立編輯列
                cols = st.columns(len(week_data))
                
                for i, day_data in enumerate(week_data):
                    with cols[i]:
                        day = day_data['day']
                        
                        # 取得原始班次（從原始資料庫中）
                        original_shift = ""
                        for col_idx in matching_columns:
                            column_data = df.iloc[:, col_idx]
                            row_idx = day + 2
                            
                            if row_idx < len(column_data):
                                value = column_data.iloc[row_idx]
                                
                                # 更嚴格的空值檢查
                                if pd.isna(value) or value is None:
                                    continue
                                    
                                shift_value = str(value).strip()
                                
                                if shift_value and shift_value.lower() not in ['nan', 'none', '']:
                                    original_shift = shift_value
                                    break
                        
                        # 取得目前有效的班次（可能是手動修改過的）
                        effective_shift = DataProcessor.get_effective_shift(
                            df, preview_data.personnel, preview_data.year, preview_data.month, day, matching_columns
                        )
                        
                        # 處理顯示用的班次（空班次顯示為空，而不是"休假"）
                        display_shift = effective_shift if effective_shift else ""
                        
                        # 顯示日期和星期
                        st.write(f"**{day_data['日期']}**")
                        st.caption(f"{day_data['星期']}")
                        
                        # 班次選擇框
                        try:
                            default_index = shift_options.index(display_shift) if display_shift in shift_options else 0
                        except ValueError:
                            default_index = 0
                        
                        # 建立唯一的key，並檢查是否已經有這個key的值
                        selectbox_key = f"shift_edit_{preview_data.personnel}_{preview_data.year}_{preview_data.month}_{day}"
                        
                        # 處理 session state 中的 selectbox 值
                        if selectbox_key not in st.session_state:
                            # 首次初始化，設定為當前有效班次
                            st.session_state[selectbox_key] = display_shift
                        
                        new_shift = st.selectbox(
                            f"班次",
                            shift_options,
                            index=default_index,
                            key=selectbox_key,
                            label_visibility="collapsed",
                            on_change=ShiftEditor._on_shift_change,
                            args=(preview_data.personnel, preview_data.year, preview_data.month, day, original_shift)
                        )
                        
                        # 顯示修改標記（只檢查是否真的有手動修改）
                        manual_shift = SessionStateManager.get_manual_shift(
                            preview_data.personnel, preview_data.year, preview_data.month, day
                        )
                        if manual_shift is not None:
                            # 進一步檢查手動設定的值是否真的與原始值不同
                            if manual_shift != original_shift:
                                st.caption("✏️ 已修改")
                            else:
                                # 如果手動設定的值與原始值相同，清除手動設定
                                key = SessionStateManager.get_manual_shift_key(
                                    preview_data.personnel, preview_data.year, preview_data.month
                                )
                                date_str = f"{preview_data.year}/{preview_data.month:02d}/{day:02d}"
                                if (key in st.session_state.manual_shifts and 
                                    date_str in st.session_state.manual_shifts[key]):
                                    del st.session_state.manual_shifts[key][date_str]
                
                st.markdown("---")
    
    @staticmethod
    def _on_shift_change(personnel: str, year: int, month: int, day: int, original_shift: str):
        """當班次選擇改變時的回調函數"""
        selectbox_key = f"shift_edit_{personnel}_{year}_{month}_{day}"
        
        if selectbox_key in st.session_state:
            new_shift = st.session_state[selectbox_key]
            
            # 只有當新選擇的班次與原始班次不同時，才記錄為手動修改
            if new_shift != original_shift:
                SessionStateManager.set_manual_shift(personnel, year, month, day, new_shift)
            else:
                # 如果改回原始班次，移除手動設定
                key = SessionStateManager.get_manual_shift_key(personnel, year, month)
                date_str = f"{year}/{month:02d}/{day:02d}"
                if (key in st.session_state.manual_shifts and 
                    date_str in st.session_state.manual_shifts[key]):
                    del st.session_state.manual_shifts[key][date_str]
    
    @staticmethod
    def _group_days_by_week(data: List[Dict], year: int, month: int) -> Dict[int, List[Dict]]:
        """將日期按週分組"""
        weeks = {}
        
        for day_data in data:
            day = day_data['day']
            try:
                current_date = date(year, month, day)
                # 計算週數（以月初為基準）
                month_start = date(year, month, 1)
                days_from_start = (current_date - month_start).days
                week_num = (days_from_start // 7) + 1
                
                if week_num not in weeks:
                    weeks[week_num] = []
                weeks[week_num].append(day_data)
            except ValueError:
                continue
        
        return weeks
    
    @staticmethod
    def _clear_month_modifications(personnel: str, year: int, month: int):
        """清除指定月份的所有手動修改"""
        key = SessionStateManager.get_manual_shift_key(personnel, year, month)
        if key in st.session_state.manual_shifts:
            del st.session_state.manual_shifts[key]
    
    @staticmethod
    def _render_modification_stats(preview_data: PreviewData):
        """顯示修改統計資訊"""
        key = SessionStateManager.get_manual_shift_key(preview_data.personnel, preview_data.year, preview_data.month)
        
        if key in st.session_state.manual_shifts:
            modifications = st.session_state.manual_shifts[key]
            
            if modifications:
                st.subheader("📊 修改統計")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.metric("本月修改天數", f"{len(modifications)} 天")
                
                with col2:
                    work_days = sum(1 for shift in modifications.values() if shift.strip())
                    st.metric("修改為上班", f"{work_days} 天")
                
                # 顯示修改明細
                with st.expander("📋 修改明細", expanded=False):
                    for date_str, shift in sorted(modifications.items()):
                        display_shift = shift if shift.strip() else "休假"
                        st.write(f"• {date_str}: {display_shift}")

class ExcelExporter:
    """Excel 匯出功能"""
    
    @staticmethod
    def export_to_excel(query_result: QueryResult) -> Tuple[bool, Union[io.BytesIO, str], float, float, float, int]:
        """
        導出Excel報表（支援手動修改的班次）
        
        Args:
            query_result: 查詢結果物件
            
        Returns:
            (成功標誌, 檔案內容或錯誤訊息, 平日總時數, 假日總時數, 總時數, 資料行數)
        """
        try:
            df = st.session_state.df
            shift_dict = st.session_state.shift_dict
            
            # 收集原始時間字串（考慮手動修改）
            date_time_strings = ExcelExporter._collect_time_strings_with_manual(
                df, shift_dict, query_result.matching_columns, query_result.year, query_result.month, query_result.target_personnel
            )
            
            # 建立Excel資料
            excel_data = ExcelExporter._build_excel_data(
                date_time_strings, query_result.daily_breakdown, query_result.year, query_result.month
            )
            
            # 生成Excel檔案
            output = ExcelExporter._create_excel_file(excel_data, query_result.target_personnel)
            
            # 計算統計資料
            total_weekday = sum(row['平日時數'] for row in excel_data)
            total_weekend = sum(row['假日時數'] for row in excel_data)
            total_hours = total_weekday + total_weekend
            
            return True, output, total_weekday, total_weekend, total_hours, len(excel_data)
            
        except Exception as e:
            return False, f"Excel匯出失敗: {str(e)}", 0, 0, 0, 0
    
    @staticmethod
    def _collect_time_strings_with_manual(df: pd.DataFrame, shift_dict: Dict, matching_columns: List[int], year: int, month: int, personnel: str) -> Dict[str, List[str]]:
        """收集原始時間字串（支援手動修改的班次）"""
        date_time_strings = defaultdict(list)
        
        for day in DateHelper.get_month_date_range(year, month):
            try:
                current_date = date(year, month, day)
                date_str = f"{year}/{month:02d}/{day:02d}"
                
                # 取得有效班次（優先使用手動設定）
                effective_shift = DataProcessor.get_effective_shift(
                    df, personnel, year, month, day, matching_columns
                )
                
                if effective_shift in shift_dict and effective_shift:
                    shift_info = shift_dict[effective_shift]
                    
                    # 收集當天時間字串
                    current_day_strings = []
                    
                    if pd.notna(shift_info.overtime_hours_1) and str(shift_info.overtime_hours_1).strip():
                        current_day_strings.append(str(shift_info.overtime_hours_1).strip())
                    
                    if pd.notna(shift_info.overtime_hours_2) and str(shift_info.overtime_hours_2).strip():
                        current_day_strings.append(str(shift_info.overtime_hours_2).strip())
                    
                    if current_day_strings:
                        date_time_strings[date_str].extend(current_day_strings)
                    
                    # 處理跨天時間字串
                    if pd.notna(shift_info.cross_day_hours) and str(shift_info.cross_day_hours).strip():
                        cross_day_str = str(shift_info.cross_day_hours).strip()
                        next_date = current_date + timedelta(days=1)
                        next_date_str = f"{next_date.year}/{next_date.month:02d}/{next_date.day:02d}"
                        date_time_strings[next_date_str].append(cross_day_str)
            
            except ValueError:
                continue
        
        return date_time_strings
    
    @staticmethod
    def _build_excel_data(date_time_strings: Dict[str, List[str]], daily_breakdown: Dict[str, float], year: int, month: int) -> List[Dict]:
        """建立Excel資料"""
        excel_data = []
        
        for day in DateHelper.get_month_date_range(year, month):
            try:
                current_date = date(year, month, day)
                date_str = f"{year}/{month:02d}/{day:02d}"
                day_type, is_weekend = DateHelper.get_day_type(year, month, day)
                
                time_strings = date_time_strings.get(date_str, [])
                original_time_str = ",".join(time_strings) if time_strings else ""
                
                weekday_hours = 0.0
                weekend_hours = 0.0
                
                if date_str in daily_breakdown:
                    total_hours = daily_breakdown[date_str]
                    
                    if is_weekend:
                        weekend_hours = total_hours
                        # 應用修改後的假日邏輯
                        original_time_str, weekend_hours = ExcelExporter._apply_weekend_logic(
                            original_time_str, weekend_hours
                        )
                    else:
                        weekday_hours = total_hours
                
                # 處理工作類型
                work_type = ""
                if date_str in daily_breakdown and not original_time_str:
                    original_time_str = "14:00-16:00(會議)"
                    work_type = "會議"
                else:
                    work_type = TextProcessor.extract_chinese_note(original_time_str)
                
                # 只有有資料的日期才加入
                if original_time_str or weekday_hours > 0 or weekend_hours > 0:
                    excel_data.append({
                        '日期': f"{day:02d}",
                        '原始時間字串': original_time_str,
                        '平日時數': weekday_hours,
                        '假日時數': weekend_hours,
                        '工作類型': work_type
                    })
            
            except ValueError:
                continue
        
        return excel_data
    
    @staticmethod
    def _apply_weekend_logic(original_time_str: str, weekend_hours: float) -> Tuple[str, float]:
        """應用修改後的假日加班邏輯"""
        if weekend_hours <= Config.WEEKEND_MIN_HOURS_THRESHOLD and weekend_hours > 0:
            if original_time_str:
                # 提取第一個時間的結束時間
                first_time_part = original_time_str.split(',')[0]
                if '-' in first_time_part:
                    end_time = first_time_part.split('-')[1].strip()
                    try:
                        # 解析結束時間
                        if ':' in end_time:
                            end_hour = int(end_time.split(':')[0])
                            end_minute = int(end_time.split(':')[1])
                        else:
                            end_hour = int(end_time[:2]) if len(end_time) >= 2 else int(end_time)
                            end_minute = 0
                        
                        # 判斷是在前面加還是後面加
                        if end_hour < Config.EARLY_MORNING_CUTOFF:
                            # 結束時間在05:00之前，在後面加2小時
                            new_start_hour = end_hour
                            new_start_minute = end_minute
                            new_end_hour = end_hour + 2
                            if new_end_hour >= 24:
                                new_end_hour -= 24
                            
                            new_time_part = f"{new_start_hour:02d}:{new_start_minute:02d}-{new_end_hour:02d}:{end_minute:02d}(撰寫病歷)"
                            original_time_str = original_time_str + "," + new_time_part
                        else:
                            # 結束時間在05:00之後，在前面加2小時
                            start_time = first_time_part.split('-')[0].strip()
                            if ':' in start_time:
                                start_hour = int(start_time.split(':')[0])
                                start_minute = int(start_time.split(':')[1])
                            else:
                                start_hour = int(start_time[:2]) if len(start_time) >= 2 else int(start_time)
                                start_minute = 0
                            
                            new_start_hour = start_hour - 2
                            if new_start_hour < 0:
                                new_start_hour += 24
                            
                            new_time_part = f"{new_start_hour:02d}:{start_minute:02d}-{start_hour:02d}:{start_minute:02d}(撰寫病歷)"
                            original_time_str = new_time_part + "," + original_time_str
                        
                        weekend_hours += Config.AUTO_ADD_HOURS
                        
                    except (ValueError, IndexError):
                        # 解析失敗，使用預設
                        original_time_str = "12:00-14:00(撰寫病歷)," + original_time_str
                        weekend_hours += Config.AUTO_ADD_HOURS
                else:
                    original_time_str = "12:00-14:00(撰寫病歷)," + original_time_str
                    weekend_hours += Config.AUTO_ADD_HOURS
            else:
                original_time_str = "12:00-14:00(撰寫病歷)"
                weekend_hours += Config.AUTO_ADD_HOURS
        
        return original_time_str, weekend_hours
    
    @staticmethod
    def _create_excel_file(excel_data: List[Dict], target_personnel: str) -> io.BytesIO:
        """創建Excel檔案"""
        df_excel = pd.DataFrame(excel_data)
        
        # 創建Excel內容到內存
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{target_personnel}加班統計"
        
        # 設定標題
        headers = ['日期', '原始時間字串', '平日時數', '假日時數', '工作類型']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', size=12)
        
        # 設定邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 填入資料
        for row_idx, row_data in enumerate(df_excel.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                if col_idx in [3, 4]:  # 平日時數、假日時數
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if value > 0:
                        cell.number_format = '0.0'
                elif col_idx == 5:  # 工作類型
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 調整欄寬
        column_widths = [8, 30, 12, 12, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # 添加統計
        total_weekday = df_excel['平日時數'].sum()
        total_weekend = df_excel['假日時數'].sum()
        total_hours = total_weekday + total_weekend
        
        last_row = len(df_excel) + 3
        
        ws.cell(row=last_row, column=1, value="統計總計").font = Font(bold=True, size=12)
        ws.cell(row=last_row, column=1).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        
        ws.cell(row=last_row + 1, column=1, value="平日加班總時數:")
        ws.cell(row=last_row + 1, column=2, value=f"{total_weekday:.1f} 小時")
        
        ws.cell(row=last_row + 2, column=1, value="假日加班總時數:")
        ws.cell(row=last_row + 2, column=2, value=f"{total_weekend:.1f} 小時")
        
        ws.cell(row=last_row + 3, column=1, value="總加班時數:")
        ws.cell(row=last_row + 3, column=2, value=f"{total_hours:.1f} 小時")
        ws.cell(row=last_row + 3, column=2).font = Font(bold=True)
        
        wb.save(output)
        output.seek(0)
        
        return output

# ===== 主要界面函數 =====
def main():
    """主程式入口"""
    # 初始化 Session State
    SessionStateManager.initialize()
    
    st.title("🏢 員工班表加班時數統計系統")
    st.caption("v2.2 新增手動編輯班次功能 - 指定人員專用 (修復版)")
    
    # 側邊欄
    render_sidebar()
    
    # 顯示系統狀態
    render_system_status()
    
    # 根據當前頁面顯示對應內容
    page_router()

def render_sidebar():
    """渲染側邊欄"""
    with st.sidebar:
        st.header("📋 系統功能")
        
        # 功能按鈕
        if st.button("📥 載入班表資料", type="primary" if st.session_state.current_page == "載入班表資料" else "secondary"):
            st.session_state.current_page = "載入班表資料"
            st.rerun()
        
        if st.button("🔍 查詢加班時數", type="primary" if st.session_state.current_page == "查詢加班時數" else "secondary"):
            st.session_state.current_page = "查詢加班時數"
            st.rerun()
        
        if st.button("🗓️ 自定義假日管理", type="primary" if st.session_state.current_page == "自定義假日管理" else "secondary"):
            st.session_state.current_page = "自定義假日管理"
            st.rerun()
        
        # 其他功能
        render_additional_features()
        
        # 系統資訊
        render_system_info()

def render_additional_features():
    """渲染額外功能"""
    st.markdown("---")
    st.markdown("### 📝 其他功能")
    
    # 空白加班單連結
    overtime_form_url = Config.OVERTIME_FORM_URL
    
    st.markdown(f"""
    <a href="{overtime_form_url}" target="_blank">
        <button style="
            background-color: #f0f2f6;
            border: 1px solid #ddd;
            border-radius: 4px;
            padding: 8px 16px;
            font-size: 14px;
            cursor: pointer;
            width: 100%;
            text-align: center;
            color: #262730;
        ">
            📄 開啟空白加班單
        </button>
    </a>
    """, unsafe_allow_html=True)
    
    # 顯示手動修改統計
    if st.session_state.manual_shifts:
        st.markdown("---")
        st.markdown("### ✏️ 班次修改統計")
        total_modifications = sum(len(shifts) for shifts in st.session_state.manual_shifts.values())
        st.caption(f"📊 總修改次數: {total_modifications}")
        
        if st.button("🗑️ 清除所有修改", type="secondary", help="清除所有手動修改的班次"):
            st.session_state.manual_shifts.clear()
            st.success("✅ 已清除所有班次修改")
            st.rerun()

def render_system_info():
    """渲染系統資訊"""
    st.markdown("---")
    st.markdown("### ℹ️ 系統資訊")
    
    if st.session_state.data_load_time:
        st.caption(f"⏰ 資料載入時間: {st.session_state.data_load_time.strftime('%Y-%m-%d %H:%M:%S')}")
    
    st.caption(f"🔄 快取版本: {st.session_state.cache_version}")
    
    # 清除快取按鈕
    if st.button("🗑️ 清除快取", type="secondary", help="清除所有快取資料，強制重新載入"):
        SessionStateManager.clear_cache()
        st.success("✅ 快取已清除")
        st.rerun()

def render_system_status():
    """渲染系統狀態"""
    if st.session_state.df is not None:
        personnel_count = DataValidator.count_allowed_personnel(st.session_state.df)
        
        # 顯示系統狀態資訊
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.success(f"📊 班表已載入 ({personnel_count} 人)")
        
        with col2:
            # 顯示自定義假日資訊
            if st.session_state.custom_holidays:
                current_year = datetime.now().year
                current_month = datetime.now().month
                current_month_holidays = {k: v for k, v in st.session_state.custom_holidays.items()
                                        if k.startswith(f"{current_year}-{current_month:02d}-")}
                
                if current_month_holidays:
                    st.info(f"🏖️ 自定義假日: {len(current_month_holidays)} 天")
                else:
                    st.info(f"🏖️ 自定義假日: {len(st.session_state.custom_holidays)} 天")
            else:
                st.info("📅 無自定義假日")
        
        with col3:
            # 顯示班次修改資訊
            if st.session_state.manual_shifts:
                total_modifications = sum(len(shifts) for shifts in st.session_state.manual_shifts.values())
                st.warning(f"✏️ 班次修改: {total_modifications} 次")
            else:
                st.info("📋 無班次修改")
    else:
        st.warning("📋 尚未載入任何班表資料")

def page_router():
    """頁面路由"""
    if st.session_state.current_page == "載入班表資料":
        load_data_page()
    elif st.session_state.current_page == "查詢加班時數":
        query_page()
    elif st.session_state.current_page == "自定義假日管理":
        holiday_management_page()

def load_data_page():
    """載入資料頁面"""
    st.header("📥 載入雲端班表資料")
    
    with st.form("load_data_form"):
        main_sheet_url = st.text_area(
            "員工班表 Google Sheets 連結",
            placeholder="請貼上員工班表的 Google Sheets 完整連結",
            help="請確保 Google Sheets 已設定為「知道連結的使用者」可檢視"
        )
        
        st.info("📋 班種對照表: 系統將自動使用預設的班種對照表")
        
        col1, col2 = st.columns(2)
        with col1:
            submit_button = st.form_submit_button("📥 載入雲端班表", type="primary")
        with col2:
            load_default = st.form_submit_button("🔄 載入預設班表")
    
    # 處理載入預設班表
    if load_default:
        main_sheet_url = Config.DEFAULT_MAIN_SHEET_URL
        st.info("✅ 使用預設班表連結")
        submit_button = True
    
    if submit_button:
        if not main_sheet_url.strip():
            st.error("❌ 請輸入員工班表的 Google Sheets 連結")
            return
        
        with st.spinner("🔄 正在載入班表資料..."):
            df, shift_dict, message = DataLoader.load_data_from_urls(
                main_sheet_url, st.session_state.cache_version
            )
        
        if df is not None:
            # 更新 session state
            st.session_state.df = df
            st.session_state.shift_dict = shift_dict
            SessionStateManager.clear_cache()
            
            st.success(message)
            
            # 顯示資料預覽
            with st.expander("📊 資料預覽", expanded=False):
                st.write("**班表前5行資料:**")
                st.dataframe(df.head())
                
                st.write("**班種對照表:**")
                shift_preview = []
                for shift_type, shift_info in list(shift_dict.items())[:10]:
                    shift_preview.append({
                        '班種': shift_type,
                        '加班時數1': shift_info.overtime_hours_1,
                        '加班時數2': shift_info.overtime_hours_2,
                        '跨日時數': shift_info.cross_day_hours
                    })
                st.dataframe(pd.DataFrame(shift_preview))
        else:
            st.error(message)

def query_page():
    """查詢頁面"""
    st.header("🔍 員工加班時數查詢")
    
    if st.session_state.df is None:
        st.warning("⚠️ 請先載入班表資料")
        return
    
    df = st.session_state.df
    personnel_options = DataProcessor.get_personnel_options(df)
    
    if not personnel_options:
        st.error("❌ 未找到指定的人事號")
        st.info(f"📋 系統僅支援以下人事號: {', '.join(Config.ALLOWED_PERSONNEL)}")
        return
    
    # 查詢表單
    with st.form("query_form"):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            selected_personnel = st.selectbox("選擇人事號", personnel_options)
        
        with col2:
            year = st.number_input("西元年", min_value=Config.MIN_YEAR, max_value=Config.MAX_YEAR, 
                                 value=datetime.now().year)
        
        with col3:
            month = st.selectbox("月份", 
                               [(i, f"{i}月") for i in range(1, 13)], 
                               index=datetime.now().month-1,
                               format_func=lambda x: x[1])
        
        col_query, col_preview, col_edit = st.columns(3)
        with col_query:
            submit_query = st.form_submit_button("🔍 查詢加班時數", type="primary")
        with col_preview:
            preview_schedule = st.form_submit_button("👁️ 預覽班表", type="secondary")
        with col_edit:
            edit_schedule = st.form_submit_button("✏️ 編輯班表", type="secondary")
    
    # 處理班表預覽
    if preview_schedule:
        handle_schedule_preview(selected_personnel, year, month[0], df, editable=False)
    
    # 處理班表編輯
    if edit_schedule:
        handle_schedule_preview(selected_personnel, year, month[0], df, editable=True)
        st.session_state.editing_mode = True
    
    # 顯示班表預覽或編輯
    if st.session_state.preview_data is not None:
        if st.session_state.editing_mode and st.session_state.preview_data.editable:
            ShiftEditor.render_shift_editor(st.session_state.preview_data)
        else:
            render_schedule_preview()
    
    # 處理查詢
    if submit_query:
        handle_overtime_query(selected_personnel, year, month[0], df)
    
    # Excel 匯出功能
    if st.session_state.last_query_result is not None:
        render_excel_export()

def handle_schedule_preview(selected_personnel: str, year: int, month: int, df: pd.DataFrame, editable: bool = False):
    """處理班表預覽"""
    target_personnel = selected_personnel.split(' (')[0]
    
    # 驗證參數
    is_valid, error_msg = DataValidator.validate_query_parameters(target_personnel, year, month)
    if not is_valid:
        st.error(f"❌ {error_msg}")
        return
    
    matching_columns = DataProcessor.find_matching_personnel_columns(df, target_personnel)
    
    if matching_columns:
        action_text = "編輯" if editable else "預覽"
        with st.spinner(f"👁️ 正在生成 {target_personnel} 的 {year}年{month}月 班表{action_text}..."):
            preview_data = SchedulePreview.generate_schedule_preview(
                target_personnel, year, month, matching_columns, editable
            )
            st.session_state.preview_data = preview_data
            
            if editable:
                st.session_state.current_edit_key = SessionStateManager.get_manual_shift_key(target_personnel, year, month)
    else:
        st.error(f"❌ 未找到人事號: {target_personnel}")

def render_schedule_preview():
    """渲染班表預覽"""
    preview_info = st.session_state.preview_data
    st.subheader(f"👁️ {preview_info.personnel} - {preview_info.year}年{preview_info.month}月班表預覽")
    
    # 顯示統計資訊
    total_days = len(preview_info.data)
    work_days = sum(1 for item in preview_info.data if item['班次'] != '休假')
    holiday_work_days = sum(1 for item in preview_info.data if item['班次'] != '休假' and item['類型'] == '假日')
    manual_changes = sum(1 for item in preview_info.data if item['手動修改'])
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("總天數", f"{total_days} 天")
    with col2:
        st.metric("上班天數", f"{work_days} 天")
    with col3:
        st.metric("假日上班", f"{holiday_work_days} 天")
    with col4:
        st.metric("手動修改", f"{manual_changes} 天", delta="修改" if manual_changes > 0 else None)
    
    # 顯示詳細班表
    df_preview = pd.DataFrame(preview_info.data)
    
    # 移除內部使用的欄位
    display_columns = ['日期', '星期', '班次', '類型', '手動修改']
    df_display = df_preview[display_columns]
    
    st.dataframe(df_display, use_container_width=True)
    
    # 如果有手動修改，顯示編輯按鈕
    if manual_changes > 0:
        col1, col2 = st.columns(2)
        with col1:
            if st.button("✏️ 進入編輯模式", type="secondary"):
                st.session_state.editing_mode = True
                st.session_state.preview_data.editable = True
                st.rerun()
        with col2:
            if st.button("🗑️ 清除本月修改", type="secondary"):
                ShiftEditor._clear_month_modifications(preview_info.personnel, preview_info.year, preview_info.month)
                st.success("✅ 已清除本月所有手動修改")
                st.rerun()

def handle_overtime_query(selected_personnel: str, year: int, month: int, df: pd.DataFrame):
    """處理加班時數查詢（支援手動修改的班次）"""
    target_personnel = selected_personnel.split(' (')[0]
    
    # 驗證參數
    is_valid, error_msg = DataValidator.validate_query_parameters(target_personnel, year, month)
    if not is_valid:
        st.error(f"❌ {error_msg}")
        return
    
    with st.spinner(f"🔍 正在查詢 {target_personnel} 的 {year}年{month}月 加班時數..."):
        # 查找匹配的欄位
        matching_columns = DataProcessor.find_matching_personnel_columns(df, target_personnel)
        
        if not matching_columns:
            st.error(f"❌ 未找到人事號: {target_personnel}")
            return
        
        # 計算加班時數（會自動使用手動修改的班次）
        query_result = OvertimeCalculator.calculate_overtime_summary(
            target_personnel, year, month, matching_columns
        )
        
        # 儲存查詢結果
        st.session_state.last_query_result = query_result
    
    # 顯示查詢結果
    render_query_results(query_result)

def render_query_results(query_result: QueryResult):
    """渲染查詢結果"""
    st.success("✅ 查詢完成！")
    
    # 檢查是否使用了手動修改的班次
    manual_key = SessionStateManager.get_manual_shift_key(query_result.target_personnel, query_result.year, query_result.month)
    if manual_key in st.session_state.manual_shifts and st.session_state.manual_shifts[manual_key]:
        manual_count = len(st.session_state.manual_shifts[manual_key])
        st.info(f"ℹ️ 本次查詢使用了 {manual_count} 天手動修改的班次資料")
    
    # 顯示自定義假日資訊
    render_custom_holidays_info(query_result.year, query_result.month)
    
    # 統計結果卡片
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("平日加班時數", f"{query_result.weekday_hours:.1f} 小時", 
                 delta=f"{query_result.weekday_hours - Config.MAX_WEEKDAY_HOURS:.1f}" if query_result.weekday_hours != Config.MAX_WEEKDAY_HOURS else None)
    with col2:
        st.metric("假日加班時數", f"{query_result.weekend_hours:.1f} 小時")
    with col3:
        st.metric("總加班時數", f"{query_result.total_hours:.1f} 小時")
    
    # 詳細每日資料
    if query_result.daily_breakdown:
        render_daily_breakdown(query_result.daily_breakdown, query_result.year, query_result.month)

def render_custom_holidays_info(year: int, month: int):
    """渲染自定義假日資訊"""
    if st.session_state.custom_holidays:
        current_month_holidays = {k: v for k, v in st.session_state.custom_holidays.items()
                                if k.startswith(f"{year}-{month:02d}-")}
        if current_month_holidays:
            st.info(f"🏖️ 本月自定義假日 ({len(current_month_holidays)} 天)")
            holiday_list = []
            for date_key, desc in sorted(current_month_holidays.items()):
                holiday_list.append(f"• {date_key}: {desc}")
            st.markdown("\n".join(holiday_list))

def render_daily_breakdown(daily_breakdown: Dict[str, float], year: int, month: int):
    """渲染每日明細"""
    st.subheader("📅 詳細每日加班記錄")
    
    # 創建表格數據
    table_data = []
    for date_str, hours in sorted(daily_breakdown.items()):
        if hours > 0:
            try:
                date_parts = date_str.split('/')
                check_year = int(date_parts[0])
                check_month = int(date_parts[1])
                check_day = int(date_parts[2])
                day_type, is_weekend = DateHelper.get_day_type(check_year, check_month, check_day)
                
                table_data.append({
                    '日期': date_str,
                    '星期': day_type,
                    '加班時數': f"{hours:.1f}小時",
                    '類型': '假日' if is_weekend else '平日'
                })
            except (ValueError, IndexError):
                continue
    
    if table_data:
        df_display = pd.DataFrame(table_data)
        
        # 按類型分組顯示
        weekday_data = [row for row in table_data if row['類型'] == '平日']
        weekend_data = [row for row in table_data if row['類型'] == '假日']
        
        col1, col2 = st.columns(2)
        
        with col1:
            if weekday_data:
                st.write("**平日加班記錄:**")
                st.dataframe(pd.DataFrame(weekday_data), use_container_width=True)
        
        with col2:
            if weekend_data:
                st.write("**假日加班記錄:**")
                st.dataframe(pd.DataFrame(weekend_data), use_container_width=True)

def render_excel_export():
    """渲染Excel匯出功能"""
    st.subheader("📊 匯出報表")
    
    result = st.session_state.last_query_result
    
    col1, col2 = st.columns([3, 1])
    
    with col1:
        # 顯示匯出資訊，包含手動修改提示
        manual_key = SessionStateManager.get_manual_shift_key(result.target_personnel, result.year, result.month)
        if manual_key in st.session_state.manual_shifts and st.session_state.manual_shifts[manual_key]:
            manual_count = len(st.session_state.manual_shifts[manual_key])
            st.info(f"📋 準備匯出: {result.target_personnel} - {result.year}年{result.month:02d}月加班統計 (含 {manual_count} 天手動修改)")
        else:
            st.info(f"📋 準備匯出: {result.target_personnel} - {result.year}年{result.month:02d}月加班統計")
    
    with col2:
        export_button = st.button("📊 產生Excel報表", type="secondary", key="export_excel_btn")
    
    if export_button:
        with st.spinner("📊 正在產生Excel報表..."):
            success, file_content_or_error, weekday_total, weekend_total, total_hours_export, row_count = ExcelExporter.export_to_excel(result)
            
            if success:
                filename = f"{result.target_personnel}_{result.year}年{result.month:02d}月_加班時數統計.xlsx"
                
                st.success("✅ Excel報表產生成功！")
                
                # 顯示統計資訊
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("平日時數", f"{weekday_total:.1f}h")
                with col2:
                    st.metric("假日時數", f"{weekend_total:.1f}h")
                with col3:
                    st.metric("資料筆數", f"{row_count}筆")
                
                # 提供下載按鈕
                st.download_button(
                    label="📥 下載Excel檔案",
                    data=file_content_or_error.getvalue(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officeedocument.spreadsheetml.sheet",
                    key="download_excel_btn"
                )
            else:
                st.error(f"❌ {file_content_or_error}")

def holiday_management_page():
    """自定義假日管理頁面"""
    st.header("🗓️ 自定義假日管理")
    
    # 提醒使用者沒有記憶功能
    st.warning("⚠️ 注意：自定義假日設定在關閉瀏覽器或重新載入頁面後將會清除，沒有記憶功能。")
    
    # 新增假日區域
    render_add_holiday_form()
    
    # 管理現有假日
    render_existing_holidays()

def render_add_holiday_form():
    """渲染新增假日表單"""
    st.subheader("➕ 新增自定義假日")
    
    with st.form("add_holiday_form"):
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            holiday_year = st.number_input("年份", min_value=Config.MIN_YEAR, max_value=Config.MAX_YEAR, 
                                         value=datetime.now().year)
        
        with col2:
            holiday_month = st.selectbox("月份", 
                                       [(i, f"{i}月") for i in range(1, 13)], 
                                       index=datetime.now().month-1,
                                       format_func=lambda x: x[1])
        
        with col3:
            holiday_day = st.number_input("日期", min_value=1, max_value=31, value=1)
        
        with col4:
            holiday_reason = st.text_input("假日原因", value="自定義假日")
        
        col_add, col_remove = st.columns(2)
        with col_add:
            add_holiday = st.form_submit_button("➕ 新增假日", type="primary")
        with col_remove:
            remove_holiday = st.form_submit_button("❌ 移除假日", type="secondary")
    
    # 處理新增假日
    if add_holiday:
        year_val = holiday_year
        month_val = holiday_month[0]
        day_val = holiday_day
        reason = holiday_reason.strip() if holiday_reason.strip() else "自定義假日"
        
        add_holiday_to_session(year_val, month_val, day_val, reason)
    
    # 處理移除假日
    if remove_holiday:
        year_val = holiday_year
        month_val = holiday_month[0]
        day_val = holiday_day
        date_key = f"{year_val}-{month_val:02d}-{day_val:02d}"
        
        if date_key in st.session_state.custom_holidays:
            removed = st.session_state.custom_holidays.pop(date_key)
            st.success(f"✅ 已移除自定義假日: {date_key} ({removed})")
            st.rerun()
        else:
            st.warning(f"⚠️ 該日期不是自定義假日: {date_key}")

def add_holiday_to_session(year: int, month: int, day: int, reason: str):
    """添加假日到session"""
    try:
        test_date = date(year, month, day)
        date_key = f"{year}-{month:02d}-{day:02d}"
        
        weekdays = ['一', '二', '三', '四', '五', '六', '日']
        weekday = weekdays[test_date.weekday()]
        
        st.session_state.custom_holidays[date_key] = f"{reason}({weekday})"
        st.success(f"✅ 已新增假日: {date_key} {reason}({weekday})")
        st.rerun()
    except ValueError:
        st.error(f"❌ 無效日期: {year}-{month:02d}-{day:02d}")

def render_existing_holidays():
    """渲染現有假日管理"""
    st.subheader("📅 目前設定的自定義假日")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🗑️ 清除所有假日", type="secondary"):
            if st.session_state.custom_holidays:
                st.session_state.custom_holidays.clear()
                st.success("✅ 已清除所有自定義假日")
                st.rerun()
            else:
                st.info("📅 目前沒有設定任何自定義假日")
    
    with col2:
        # 匯出假日設定
        if st.session_state.custom_holidays and st.button("📄 匯出假日清單", type="secondary"):
            holiday_text = "\n".join([f"{date_key}: {desc}" for date_key, desc in sorted(st.session_state.custom_holidays.items())])
            st.download_button(
                label="📥 下載假日清單",
                data=holiday_text,
                file_name=f"自定義假日_{datetime.now().strftime('%Y%m%d')}.txt",
                mime="text/plain"
            )
    
    # 顯示假日清單
    if st.session_state.custom_holidays:
        st.write(f"**目前共有 {len(st.session_state.custom_holidays)} 個自定義假日:**")
        
        # 轉換為表格顯示
        holiday_data = []
        for date_key, description in sorted(st.session_state.custom_holidays.items()):
            holiday_data.append({
                '日期': date_key,
                '描述': description,
                '年': date_key.split('-')[0],
                '月': date_key.split('-')[1],
                '日': date_key.split('-')[2]
            })
        
        df_holidays = pd.DataFrame(holiday_data)
        
        # 分頁顯示
        if len(df_holidays) > 10:
            # 使用分頁
            page_size = 10
            total_pages = (len(df_holidays) - 1) // page_size + 1
            
            page_num = st.selectbox("選擇頁面", range(1, total_pages + 1), format_func=lambda x: f"第 {x} 頁")
            
            start_idx = (page_num - 1) * page_size
            end_idx = start_idx + page_size
            
            st.dataframe(df_holidays[['日期', '描述']].iloc[start_idx:end_idx], use_container_width=True)
            
            st.caption(f"顯示第 {start_idx + 1}-{min(end_idx, len(df_holidays))} 筆，共 {len(df_holidays)} 筆")
        else:
            st.dataframe(df_holidays[['日期', '描述']], use_container_width=True)
        
        # 按月份分組顯示
        render_holidays_by_month(df_holidays)
    else:
        st.info("📅 目前沒有設定任何自定義假日")

def render_holidays_by_month(df_holidays: pd.DataFrame):
    """按月份分組顯示假日"""
    st.subheader("📊 按月份分組")
    
    months_dict = {}
    for _, row in df_holidays.iterrows():
        year_month = row['日期'][:7]  # YYYY-MM
        if year_month not in months_dict:
            months_dict[year_month] = []
        months_dict[year_month].append(f"{row['日期']}: {row['描述']}")
    
    # 使用tabs顯示不同月份
    if months_dict:
        month_tabs = st.tabs([f"📅 {ym} ({len(holidays)}天)" for ym, holidays in sorted(months_dict.items())])
        
        for i, (year_month, holidays) in enumerate(sorted(months_dict.items())):
            with month_tabs[i]:
                for holiday in holidays:
                    st.write(f"• {holiday}")

# ===== 程式入口點 =====
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        st.error(f"❌ 系統發生錯誤: {str(e)}")
        st.info("🔄 請嘗試重新載入頁面或清除快取")
        
        # 錯誤詳情（開發模式）
        with st.expander("🔍 錯誤詳情 (開發模式)", expanded=False):
            st.exception(e)